#!/usr/bin/env python3
"""Export daily operations roster to data/operations_roster.json.

Reads xpbi02 PersonnelRosterView (normalised by normalise_rapidcrews_workbook.py)
and xll01 Personnel from the macro workbook.  Filters to Kwinana-area clients,
groups per-day rows into date-range assignments per (PersonnelId, JobNo), and
writes data/operations_roster.json for the frontend operations roster view.
"""
from __future__ import annotations

import datetime as dt
import json
import pathlib
import re
from typing import Any

import openpyxl

REPO_ROOT = pathlib.Path(__file__).resolve().parent.parent
DATA_DIR = REPO_ROOT / "data"
RAW_FILE = DATA_DIR / "raw" / "Rapidcrews Macro Data.xlsx"
OUT_FILE = DATA_DIR / "operations_roster.json"

ROSTER_VIEW_SHEET = "xpbi02 PersonnelRosterView"
PERSONNEL_SHEET = "xll01 Personnel"

# Match against combined "client site" string (lowercased).
KWINANA_TERMS = (
    "covalent", "tronox", "csbp", "tianqi",
    "kleenheat", "kwinana", "naan", "kpf",
)

ALIASES: dict[str, list[str]] = {
    # PersonnelRosterView columns (as written by normalise_rapidcrews_workbook.py)
    "Job No": ["Job No", "JobNo", "Job Number", "JobId", "Job Id", "QuoteNo", "OrderNo"],
    "Client": ["Client", "ClientId", "Client Id", "ClientName", "Client Name"],
    "Site": ["Site", "SiteName", "Site Name", "Location"],
    "Personnel Id": ["Personnel Id", "PersonnelId", "Personnel ID", "Employee Id", "EmployeeID"],
    "Schedule Date": ["Schedule Date", "ReportDate", "Date", "Roster Date", "Shift Date"],
    "Schedule Type": ["Schedule Type", "ScheduleType", "Shift", "Shift Type"],
    "IsOnLocation": ["IsOnLocation", "OnSite", "On Site", "Is On Location"],
    # xll01 Personnel columns
    "Given Names": ["Given Names", "Given Name", "First Name", "FirstName"],
    "Surname": ["Surname", "Last Name", "Family Name"],
    "Role": ["Trade", "Discipline", "Primary Role", "Role", "Position"],
    "Mobile": ["Mobile", "Mobile Phone", "Mobile No", "Phone", "Contact"],
    "Hire Company": ["Hire Company", "Hiring Company", "Labour Hire Company", "Labour Hire", "Company"],
}


def _clean(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\xa0", " ")).strip()


def _norm(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", _clean(value).lower())


def _headers(ws) -> dict[str, int]:
    try:
        header = list(next(ws.iter_rows(max_row=1, values_only=True)))
    except StopIteration:
        return {}
    return {_clean(h): i for i, h in enumerate(header) if _clean(h)}


def _find_col(headers: dict[str, int], canonical: str, required: bool = True) -> int | None:
    normal = {_norm(k): v for k, v in headers.items()}
    for alias in ALIASES.get(canonical, [canonical]):
        if _norm(alias) in normal:
            return normal[_norm(alias)]
    if required:
        raise KeyError(f"Missing required column {canonical!r}; available: {', '.join(headers)}")
    return None


def _get(row: tuple[Any, ...], idx: int | None) -> Any:
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _date(value: Any) -> str:
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    text = _clean(value)
    if not text:
        return ""
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return dt.datetime.strptime(text[:10], fmt).date().isoformat()
        except ValueError:
            pass
    return ""


def _contiguous_ranges_with_sched(
    tuples: list[tuple[str, str, bool]],
) -> list[dict[str, Any]]:
    """Split (date_str, sched_type, is_on_location) tuples into contiguous segments.

    A new segment starts when the calendar gap is > 1 day, the schedule_type
    changes, or is_on_location changes.  This ensures that off-site days (R&R,
    demob) form separate segments with is_on_location=False rather than being
    merged with preceding on-site days.
    """
    if not tuples:
        return []
    segs: list[dict[str, Any]] = []
    seg_start, seg_sched, seg_onsite = tuples[0]
    prev_date = dt.date.fromisoformat(tuples[0][0])

    for date_str, sched, onsite in tuples[1:]:
        d = dt.date.fromisoformat(date_str)
        if (d - prev_date).days > 1 or sched != seg_sched or onsite != seg_onsite:
            segs.append({
                "start": seg_start,
                "end": prev_date.isoformat(),
                "schedule_type": seg_sched,
                "is_on_location": seg_onsite,
            })
            seg_start  = date_str
            seg_sched  = sched
            seg_onsite = onsite
        prev_date = d

    segs.append({
        "start": seg_start,
        "end": prev_date.isoformat(),
        "schedule_type": seg_sched,
        "is_on_location": seg_onsite,
    })
    return segs


def _parse_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    text = _clean(value).lower()
    return text in {"1", "true", "yes", "y", "on", "onsite", "on site"}


def _resolve_sheet(wb: openpyxl.Workbook, canonical: str) -> str | None:
    if canonical in wb.sheetnames:
        return canonical
    target = _norm(canonical)
    for name in wb.sheetnames:
        n = _norm(name)
        if n == target or re.fullmatch(rf"{re.escape(target)}\d+", n):
            return name
    return None


def _is_kwinana(client: str, site: str) -> bool:
    combined = f"{client} {site}".lower()
    return any(term in combined for term in KWINANA_TERMS)


def _load_personnel(wb: openpyxl.Workbook) -> dict[str, dict[str, str]]:
    sheet = _resolve_sheet(wb, PERSONNEL_SHEET)
    if not sheet:
        print("export_operations_roster: xll01 Personnel sheet not found")
        return {}
    ws = wb[sheet]
    headers = _headers(ws)
    pid_i    = _find_col(headers, "Personnel Id", required=False)
    first_i  = _find_col(headers, "Given Names", required=False)
    last_i   = _find_col(headers, "Surname", required=False)
    role_i   = _find_col(headers, "Role", required=False)
    mobile_i = _find_col(headers, "Mobile", required=False)
    hire_i   = _find_col(headers, "Hire Company", required=False)

    out: dict[str, dict[str, str]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        pid = _clean(_get(row, pid_i))
        if not pid:
            continue
        first = _clean(_get(row, first_i))
        last  = _clean(_get(row, last_i))
        name  = f"{first} {last}".strip()
        out[pid] = {
            "name":         name,
            "role":         _clean(_get(row, role_i)),
            "mobile":       _clean(_get(row, mobile_i)),
            "hire_company": _clean(_get(row, hire_i)),
        }
    return out


def _load_roster_assignments(wb: openpyxl.Workbook) -> list[dict[str, Any]]:
    """Read PersonnelRosterView, filter to Kwinana, group per (pid, job)."""
    sheet = _resolve_sheet(wb, ROSTER_VIEW_SHEET)
    if not sheet:
        print("export_operations_roster: xpbi02 PersonnelRosterView sheet not found")
        return []

    ws = wb[sheet]
    headers = _headers(ws)
    job_i    = _find_col(headers, "Job No", required=False)
    client_i = _find_col(headers, "Client", required=False)
    site_i   = _find_col(headers, "Site", required=False)
    pid_i    = _find_col(headers, "Personnel Id", required=False)
    date_i   = _find_col(headers, "Schedule Date", required=False)
    sched_i  = _find_col(headers, "Schedule Type", required=False)
    onsite_i = _find_col(headers, "IsOnLocation", required=False)

    if date_i is None or pid_i is None:
        print("export_operations_roster: missing Schedule Date or Personnel Id column; skipped")
        return []

    # Group per-day rows into (pid, job_no) buckets.
    groups: dict[tuple[str, str], dict[str, Any]] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not any(row):
            continue
        pid = _clean(_get(row, pid_i))
        day = _date(_get(row, date_i))
        if not pid or not day:
            continue
        client = _clean(_get(row, client_i)) if client_i is not None else ""
        site   = _clean(_get(row, site_i))   if site_i   is not None else ""
        if not _is_kwinana(client, site):
            continue
        job    = _clean(_get(row, job_i))    if job_i    is not None else ""
        sched  = _clean(_get(row, sched_i))  if sched_i  is not None else ""
        onsite = _parse_bool(_get(row, onsite_i))

        if not sched:
            sched = "Day Shift"
        key = (pid, job)
        if key not in groups:
            groups[key] = {
                "pid": pid, "job": job,
                "client": client, "site": site,
                "days": {},  # date_str -> (sched_type, is_on_location)
            }
        g = groups[key]
        # When multiple rows exist for the same day, prefer is_on_location=True.
        existing = g["days"].get(day)
        if existing is None or (onsite and not existing[1]):
            g["days"][day] = (sched, onsite)

    assignments: list[dict[str, Any]] = []
    for g in groups.values():
        if not g["days"]:
            continue
        sorted_tuples = [
            (d, sc, ol)
            for d, (sc, ol) in sorted(g["days"].items())
        ]
        for seg in _contiguous_ranges_with_sched(sorted_tuples):
            assignments.append({
                "pid":            g["pid"],
                "job":            g["job"],
                "client":         g["client"],
                "site":           g["site"],
                "start":          seg["start"],
                "end":            seg["end"],
                "schedule_type":  seg["schedule_type"],
                "is_on_location": seg["is_on_location"],
            })
    return assignments


def main() -> int:
    if not RAW_FILE.exists():
        print("export_operations_roster: macro workbook not found; skipped")
        return 0

    wb = openpyxl.load_workbook(RAW_FILE, data_only=True, read_only=True)
    try:
        personnel   = _load_personnel(wb)
        assignments = _load_roster_assignments(wb)
    finally:
        wb.close()

    if not assignments:
        print("export_operations_roster: no Kwinana roster rows found; skipped")
        return 0

    # Build worker records, one entry per personnel_id.
    by_pid: dict[str, dict[str, Any]] = {}
    for a in assignments:
        pid = a["pid"]
        if pid not in by_pid:
            p = personnel.get(pid, {})
            by_pid[pid] = {
                "personnel_id": pid,
                "name":         p.get("name", ""),
                "role":         p.get("role", ""),
                "mobile":       p.get("mobile", ""),
                "hire_company": p.get("hire_company", ""),
                "assignments":  [],
            }
        by_pid[pid]["assignments"].append({
            "job_no":          a["job"],
            "client":          a["client"],
            "site":            a["site"],
            "start":           a["start"],
            "end":             a["end"],
            "schedule_type":   a["schedule_type"],
            "is_on_location":  a["is_on_location"],
        })

    workers = sorted(by_pid.values(), key=lambda w: w["name"])

    payload = {
        "generated_at": dt.datetime.utcnow().replace(microsecond=0).isoformat() + "Z",
        "worker_count": len(workers),
        "workers":      workers,
    }
    OUT_FILE.write_text(json.dumps(payload, indent=2))
    print(
        f"export_operations_roster: wrote {len(workers)} workers "
        f"({len(assignments)} assignments) to {OUT_FILE.relative_to(REPO_ROOT)}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
