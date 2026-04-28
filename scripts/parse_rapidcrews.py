#!/usr/bin/env python3
"""Convert roster XLSX exports into the dashboard's per-company JSON.

Two roster flavours are supported:

1. **Rapid Crews "RosterCut" export** — filename `<roster_id> (RosterCut) <ts>.xlsx`.
   The canonical export used for the three live shutdowns (Covalent, Tronox, CSBP).
2. **Kleenheat-style spreadsheet** — a looser, older format with first-name-only
   entries, used as a historical seed for retention/carry-over stats. Surnames
   are reconstructed by cross-referencing the first-name + role against the
   other rosters (so e.g. Kleenheat's "Joe, Intermediate Rigger" resolves to
   "Joe DACK" because Covalent has exactly one Joe DACK as Intermediate Rigger).

Workflow
--------
1. Drop the XLSX file into `data/raw/`.
2. Register it in ROSTER_MAP below (key = first filename token before the space).
3. Run `python3 scripts/parse_rapidcrews.py`.
4. Commit the regenerated `data/<company>.json` files.

End-user workflow (no code / no git)
------------------------------------
Edit the `ACTIVE_SHUTDOWNS` sheet inside `Rapidcrews Macro Data.xlsx` — add
or remove JobNo rows to control which shutdowns the dashboard shows. The
macro-data loader (`scripts/parse_macro_data.py`) is called from `main()`
below; see that module's docstring for full detail.

Data NOT in the Rapid Crews export
----------------------------------
- **Required headcount per role** (the original target on the request). The
  parser writes `required_by_role = filled_by_role` as a placeholder; override
  with real targets at `data/targets/<shutdown_id>.json`, which
  `scripts/sync_source_targets.py` populates from each site's own dashboard repo.
- **Client name** (Covalent / Tronox / CSBP / Kleenheat). The "Company" column
  inside Rapid Crews is the labour-hire firm (SRG South West, MMFS, …), not
  the client whose plant the shutdown is for — that mapping lives in
  ROSTER_MAP.
"""
from __future__ import annotations

import datetime as dt
import json
import pathlib
import re
import sys

import openpyxl


# Each ROSTER_MAP entry is (company_key, client_display_name, project_label,
# site). An optional 5th element — shutdown_id_override — lets us
# disambiguate multiple shutdowns at the same client in the same month
# (e.g. Tianqi running a Construction Ramp-Up and a Scaffold Shutdown in
# parallel through April-June 2026). Without the override, shutdown_id
# defaults to "<company_key>-<YYYY-MM of start_date>" and the two would
# collide.
#
# File-key conventions:
#   - Numeric leading token  -> Rapid Crews roster id (from the RosterCut
#                                filename convention)
#   - Anything else          -> the whole filename stem (minus .xlsx,
#                                whitespace-trimmed)
ROSTER_MAP: dict[str, tuple] = {
    # Rapid Crews RosterCut exports — keyed by numeric roster_id.
    "1353": ("tronox",   "Tronox",   "Major Shutdown May 2026", "Kwinana"),
    "1359": ("covalent", "Covalent", "Mt Holland April 2026",   "Mt Holland"),
    "1375": ("csbp",     "CSBP",     "NAAN2 June 2026",         "Kwinana"),

    # Historic Pegasus-format rosters — kept in data/raw/ so the retention
    # matrix still has pre-2026 shutdowns even after Rapid Crews' live SQL
    # view rolls these JobNos off. Dates come from the roster rows themselves
    # (Date In / Date Out). See parse_pegasus_roster().
    "1110": ("covalent", "Covalent", "Mt Holland October 2025",  "Mt Holland",
             "covalent-2025-10"),
    "1116": ("tronox",   "Tronox",   "Major Shutdown November 2025", "Kwinana",
             "tronox-2025-11"),
    "1147": ("csbp",     "CSBP",     "NAAN3 November 2025",       "Kwinana",
             "csbp-2025-11"),

    # CSBP is the umbrella WesCEF client — their Kwinana estate covers the
    # KPF LNG (Kleenheat-branded) plant and the NAAN2 fertiliser unit. Both
    # roll up under the same company_key so the unified dashboard treats
    # them as one client for filter/retention purposes. The shutdown_id for
    # the historical KPF LNG roster stays "kleenheat-2026-03" so the
    # existing target file (data/targets/kleenheat-2026-03.json) keeps
    # working without a rename.
    "Kleenheat Major March 2026":
        ("csbp", "CSBP", "KPF LNG Major Shutdown March 2026 (Kleenheat)", "Kwinana",
         "kleenheat-2026-03"),

    # Tianqi removed from active tracking at user's request (2026-04-15).
    # XLSX is still in data/raw/ — re-add the entry below to reinstate.
    # "Tianqi Construction Ramp Up Project":
    #     ("tianqi", "Tianqi", "Construction Ramp Up Project", "Kwinana",
    #      "tianqi-construction-2026-04"),
}

REPO_ROOT      = pathlib.Path(__file__).resolve().parent.parent
RAW_DIR        = REPO_ROOT / "data" / "raw"
DATA_DIR       = REPO_ROOT / "data"
TARGETS_DIR    = DATA_DIR / "targets"     # optional override: targets/<shutdown_id>.json
HISTORY_DIR    = DATA_DIR / "history"     # per-shutdown snapshot that persists
                                          # after Rapid Crews' SQL view rolls over
ENRICHMENT_DIR = DATA_DIR / "enrichment"  # per-company resume/annotation overlay
                                          # (trade years, resume prose, newhire flags —
                                          # fields not carried by the Rapid Crews feed)

RAPIDCREWS_COLS = ["Company", "Name", "Surname", "Position", "Position On Project",
                   "Start Date", "End Date", "Confirmed", "Crew Type", "Mobilised"]
KLEENHEAT_COLS  = ["Name", "Trade", "Company", "On Site", "Off Site", "Crew"]
PEGASUS_COLS    = ["Company", "Date In", "Date Out", "Shift", "Surname", "First Name",
                   "Pegasus Job Role"]


# --------------------------------------------------------------------------- helpers

def to_iso(d) -> str | None:
    if isinstance(d, dt.datetime):
        return d.date().isoformat()
    if isinstance(d, dt.date):
        return d.isoformat()
    if isinstance(d, str):
        # Spreadsheets sometimes store ISO-ish strings with a trailing 'Z'
        m = re.match(r"(\d{4}-\d{2}-\d{2})", d)
        return m.group(1) if m else None
    return None


def truthy(v) -> bool:
    return str(v or "").strip().upper() in {"YES", "Y", "TRUE", "1"}


def _standardise_mobile(raw) -> str:
    """Normalise Australian mobile numbers to canonical '04XX XXX XXX' form.

    The rosters come from three different XLSX schemas and the team enter
    mobiles inconsistently: some are local-style with sensible spacing
    ('0493 038 522'), some are mis-spaced ('049 759 4673' — should group as
    0497 594 673), some are international without a plus ('61420397028'), and
    a handful carry stray dashes or parentheses. Strip every non-digit, then
    re-pad/re-space to a single canonical form so the matrix renders cleanly
    and the tel: links dial correctly on both desktop and mobile.

    Non-mobile or unrecognisable input is returned empty — better to hide a
    dodgy number than to dial the wrong person.
    """
    if raw is None:
        return ""
    digits = re.sub(r"\D", "", str(raw))
    if not digits:
        return ""
    # International with country code: 61 4XX XXX XXX (11 digits total, leading 61)
    if len(digits) == 11 and digits.startswith("61") and digits[2] == "4":
        digits = "0" + digits[2:]
    # Nine-digit local (missing leading 0): 4XX XXX XXX -> 04XX XXX XXX
    elif len(digits) == 9 and digits.startswith("4"):
        digits = "0" + digits
    # Anything else that doesn't conform to the 10-digit 04XX mobile pattern
    # is kept as raw digits (not formatted) — unusual but preserved for ops
    # review rather than silently dropped.
    if len(digits) == 10 and digits.startswith("04"):
        return f"{digits[0:4]} {digits[4:7]} {digits[7:10]}"
    return digits


# --------------------------------------------------------------------------- roster parsers

def _detect_format(headers: list) -> str:
    """Detect which of our three supported roster schemas the XLSX is using."""
    hs = {h for h in headers if h}
    if set(RAPIDCREWS_COLS).issubset(hs):
        return "rapidcrews"
    if set(PEGASUS_COLS).issubset(hs):
        return "pegasus"
    if set(KLEENHEAT_COLS).issubset(hs):
        return "kleenheat"
    return "unknown"


def parse_rapidcrews_roster(xlsx_path: pathlib.Path) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}
    mobile_col = idx.get("Mobile")

    rows: list[dict] = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        if not any(raw):
            continue
        first = str(raw[idx["Name"]]    or "").strip()
        last  = str(raw[idx["Surname"]] or "").strip()
        name  = f"{first} {last}".strip()
        if not name:
            continue
        role = raw[idx["Position On Project"]] or raw[idx["Position"]] or "Unknown"
        mobile = _standardise_mobile(raw[mobile_col]) if mobile_col is not None else ""
        rows.append({
            "labour_hire": (raw[idx["Company"]] or "").strip(),
            "name":        name,
            "first_name":  first,
            "last_name":   last,
            "role":        str(role).strip(),
            "mobile":      mobile,
            "start":       to_iso(raw[idx["Start Date"]]),
            "end":         to_iso(raw[idx["End Date"]]),
            "confirmed":   truthy(raw[idx["Confirmed"]]),
            "crew_type":   (raw[idx["Crew Type"]] or "Unknown").strip(),
            "mobilised":   truthy(raw[idx["Mobilised"]]),
        })
    return rows


_LAST_NAME_COL_CANDIDATES = ("Last Dna", "Last Name", "Surname", "Last")


def _surname_from_email(email: str | None, first_name: str) -> str:
    """Best-effort surname extraction from an email local-part. Handles the two
    common patterns we see in the Kleenheat export:
      - `firstname.surname@...`  -> "surname"
      - `surname_first@...`,  `firstsurname@...`,  `firstsurname12@...`
        (strip digits/punct, then strip the first-name prefix/suffix if present).
    Returns "" if nothing reasonable can be derived.
    """
    if not email or "@" not in email:
        return ""
    local = email.split("@", 1)[0].lower()
    local = re.sub(r"[^a-z._]+", "", local)     # drop digits, plus, etc.
    if "." in local:
        parts = [p for p in local.split(".") if p]
        first_low = first_name.lower()
        # drop the piece that matches the first name (anywhere in the split)
        parts = [p for p in parts if p != first_low]
        if parts:
            return parts[-1].capitalize()
    # no dot: try stripping the first-name as a prefix or suffix
    flat = local.replace("_", "")
    first_low = first_name.lower()
    if flat.startswith(first_low) and len(flat) > len(first_low):
        return flat[len(first_low):].capitalize()
    if flat.endswith(first_low) and len(flat) > len(first_low):
        return flat[:-len(first_low)].capitalize()
    return ""


def parse_kleenheat_roster(xlsx_path: pathlib.Path) -> list[dict]:
    """Looser spreadsheet format: first-name-only entries, no Confirmed column.
    Every row present in the spreadsheet is treated as confirmed + mobilised
    (the shutdown has already happened).

    Surnames can come from three places, in priority order:
      1. An explicit surname column (Last Dna / Last Name / Surname / Last)
      2. The email column's local-part (firstname.surname@... style)
      3. Cross-reference against the other rosters — applied later in
         `enrich_kleenheat_names`.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}
    surname_col = next((idx[c] for c in _LAST_NAME_COL_CANDIDATES if c in idx), None)
    email_col   = idx.get("Email")
    mobile_col  = idx.get("Mobile")

    rows: list[dict] = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        if not any(raw):
            continue
        first = (raw[idx["Name"]] or "").strip()
        if not first:
            continue
        surname = ""
        resolution = None
        if surname_col is not None and raw[surname_col]:
            surname = str(raw[surname_col]).strip()
            if surname:
                resolution = "explicit_column"
        if not surname and email_col is not None:
            surname = _surname_from_email(raw[email_col], first)
            if surname:
                resolution = "email_heuristic"
        name = f"{first} {surname}".strip()
        role = str(raw[idx["Trade"]] or "Unknown").strip()
        crew = str(raw[idx["Crew"]] or "Unknown").strip().title()  # "DAY" -> "Day"
        mobile = _standardise_mobile(raw[mobile_col]) if mobile_col is not None else ""
        rows.append({
            "labour_hire":       (raw[idx["Company"]] or "").strip(),
            "name":              name,
            "first_name":        first,
            "last_name":         surname,
            "role":              role,
            "mobile":            mobile,
            "start":             to_iso(raw[idx["On Site"]]),
            "end":               to_iso(raw[idx["Off Site"]]),
            "confirmed":         True,
            "crew_type":         crew,
            "mobilised":         True,
            "_name_resolution":  resolution,      # None until enrichment fills it
        })
    return rows


def parse_pegasus_roster(xlsx_path: pathlib.Path) -> list[dict]:
    """Pegasus-style labour list: First Name + Surname columns, Date In/Out,
    Shift (DS/NS), Pegasus Job Role. Every row is treated as confirmed +
    mobilised — the shutdown has already happened."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}
    shift_label = {"DS": "Day", "NS": "Night", "DAY": "Day", "NIGHT": "Night"}
    mobile_col = idx.get("Contractor Mobile Number") or idx.get("Mobile")

    rows: list[dict] = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        if not any(raw):
            continue
        first = (raw[idx["First Name"]] or "").strip()
        last  = (raw[idx["Surname"]]    or "").strip()
        name  = f"{first} {last}".strip()
        if not name:
            continue
        role  = str(raw[idx["Pegasus Job Role"]] or "Unknown").strip()
        shift = str(raw[idx["Shift"]] or "").strip().upper()
        mobile = _standardise_mobile(raw[mobile_col]) if mobile_col is not None else ""
        rows.append({
            "labour_hire":      (raw[idx["Company"]] or "").strip(),
            "name":             name,
            "first_name":       first,
            "last_name":        last,
            "role":             role,
            "mobile":           mobile,
            "start":            to_iso(raw[idx["Date In"]]),
            "end":              to_iso(raw[idx["Date Out"]]),
            "confirmed":        True,
            "crew_type":        shift_label.get(shift, shift.title() or "Unknown"),
            "mobilised":        True,
            "_name_resolution": "explicit_column",
        })
    return rows


def parse_roster(xlsx_path: pathlib.Path) -> tuple[str, list[dict]]:
    """Sniff the XLSX, dispatch to the right parser, return (format, rows)."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb.active
    # values_only=True yields plain values, not Cell objects.
    headers = list(next(ws.iter_rows(max_row=1, values_only=True)))
    fmt = _detect_format(headers)
    wb.close()
    if fmt == "rapidcrews":
        return fmt, parse_rapidcrews_roster(xlsx_path)
    if fmt == "pegasus":
        return fmt, parse_pegasus_roster(xlsx_path)
    if fmt == "kleenheat":
        return fmt, parse_kleenheat_roster(xlsx_path)
    raise ValueError(f"{xlsx_path.name}: unrecognised roster columns {headers}")


# --------------------------------------------------------------------------- surname enrichment

def build_surname_lookup(rosters_by_company: dict[str, list[dict]]) -> dict:
    """From the other companies' rosters, build a map
        (first_name_lower, role_lower) -> set of full names
    so Kleenheat's first-name-only rows can be resolved to a full name when the
    first-name + role combination is unique across SRG's other live rosters.
    """
    lookup: dict[tuple[str, str], set[str]] = {}
    for company, rows in rosters_by_company.items():
        if company == "kleenheat":
            continue
        for r in rows:
            full = r["name"].strip()
            if " " not in full:
                continue
            first = full.split()[0].lower()
            role  = r["role"].strip().lower()
            key   = (first, role)
            lookup.setdefault(key, set()).add(full)
    return lookup


def enrich_kleenheat_names(rows: list[dict], lookup: dict) -> dict[str, int]:
    """Fill in surnames on first-name-only Kleenheat rows by cross-referencing
    first-name + role against the other rosters. Rows that already picked up a
    surname from the spreadsheet's own surname column or the email heuristic
    are left alone but still counted. Resolutions are tagged on each row via
    `_name_resolution` for the data-quality warnings panel."""
    stats = {
        "explicit_column": 0,
        "email_heuristic": 0,
        "xref_exact":      0,
        "xref_ambiguous":  0,
        "unmatched":       0,
    }
    for r in rows:
        if r.get("_name_resolution") == "explicit_column":
            stats["explicit_column"] += 1
            continue
        if r.get("_name_resolution") == "email_heuristic":
            stats["email_heuristic"] += 1
            continue
        first = r["first_name"].lower()
        role  = r["role"].strip().lower()
        candidates = lookup.get((first, role), set())
        if len(candidates) == 1:
            r["name"] = next(iter(candidates))
            r["_name_resolution"] = "xref_exact"
            stats["xref_exact"] += 1
        elif len(candidates) > 1:
            r["_name_resolution"] = f"xref_ambiguous:{len(candidates)}"
            stats["xref_ambiguous"] += 1
        else:
            r["_name_resolution"] = "unmatched"
            stats["unmatched"] += 1
    return stats


# --------------------------------------------------------------------------- targets

def merge_targets(shutdown_id: str,
                  filled_by_role: dict[str, int]
                  ) -> tuple[dict[str, int], dict[str, int], dict | None]:
    """Read optional overrides from data/targets/<shutdown_id>.json.

    Rapid Crews is the source of truth for filled_by_role — callers hand us
    the RC-derived counts and we always return them unchanged. The target
    file only contributes `required_by_role` when RC has nothing to say
    (no JobPlanningView row for this JobNo — historic Pegasus rosters, the
    Kleenheat carry-over). Any role present in the RC roster but missing
    from the target file is filled in as required = filled, so the table
    shows "0 gap" rather than an empty cell.

    Returns (required_by_role, filled_by_role, source_meta | None).
    """
    path = TARGETS_DIR / f"{shutdown_id}.json"
    if not path.exists():
        # No override: required defaults to the RC filled count (placeholder).
        return dict(filled_by_role), dict(filled_by_role), None
    data: dict = json.loads(path.read_text())

    if "required_by_role" in data:
        required_override = dict(data["required_by_role"])
        source_meta       = data.get("_source")
    else:
        # Legacy flat shape — all keys are required-by-role overrides.
        required_override = dict(data)
        source_meta       = None

    all_keys = set(filled_by_role) | set(required_override)
    required = {r: int(required_override.get(r, filled_by_role.get(r, 0)))
                for r in all_keys}
    filled   = dict(filled_by_role)
    return required, filled, source_meta


# --------------------------------------------------------------------------- build a shutdown payload

def _infer_status(start_day: dt.date, end_day: dt.date, today: dt.date) -> str:
    if end_day < today:
        return "completed"
    if start_day <= today:
        return "in_progress"
    return "booked"


def _norm_name(s) -> str:
    """Lowercase, letters-only fingerprint for name matching. Drops spaces,
    hyphens, apostrophes, diacritics so "O'Brien" collides with "OBrien"
    and "Van Der Zanden" with "VANDERZANDEN"."""
    return re.sub(r"[^a-z]+", "", (s or "").lower())


def _load_enrichment(company_key: str) -> dict[str, dict]:
    """Load data/enrichment/<company_key>.json (optional) and index every
    record by multiple normalised-name keys (first+last, last+first,
    sorted) so the feed's "Firstname SURNAME" collides with the
    enrichment file's "SURNAME, Firstname"-ish entries.

    The enrichment file carries fields the Rapid Crews feed doesn't —
    trade years, shutdown years, resume prose, newhire flags, driver's
    licence class, etc. parse_rapidcrews.py injects whatever's there
    directly onto each roster entry so per-site dashboards can stay
    thin and the consolidated feed remains single source of truth.

    Shape: {records: [{name: ..., ...overlay_fields...}, ...]}
    Returns {} when no file exists — roster is still emitted normally.
    """
    path = ENRICHMENT_DIR / f"{company_key}.json"
    if not path.exists():
        return {}
    try:
        data = json.loads(path.read_text())
    except json.JSONDecodeError as e:
        print(f"  ! enrichment {path.name}: JSON parse error: {e}", file=sys.stderr)
        return {}
    out: dict[str, dict] = {}
    for r in data.get("records", []):
        name = r.get("name", "")
        if not name:
            continue
        parts = [p for p in re.sub(r"[,]+", " ", name).split() if p]
        norm = [_norm_name(p) for p in parts if _norm_name(p)]
        if not norm:
            continue
        payload = {k: v for k, v in r.items() if k != "name"}
        keys = {"".join(norm), "".join(sorted(norm))}
        if len(norm) >= 2:
            keys.update({norm[0] + norm[-1], norm[-1] + norm[0]})
        for k in keys:
            out.setdefault(k, payload)
    return out


def _enrichment_lookup(first: str, last: str,
                       index: dict[str, dict]) -> dict | None:
    """Return the enrichment payload for (first, last), or None."""
    if not index:
        return None
    f = _norm_name(first)
    l = _norm_name(last)
    if not f or not l:
        return None
    for k in (f + l, l + f, "".join(sorted([f, l]))):
        if k in index:
            return index[k]
    return None


def _emit_roster_entries(confirmed: list[dict],
                         enrichment: dict[str, dict] | None = None) -> list[dict]:
    """Turn confirmed roster rows into the final per-worker dicts the
    dashboard reads. Enriches each entry with:
      - shift:        from RosterCut/Kleenheat crew_type when known, so the
                      per-site dashboards can group/filter without guessing
      - personnel_id: best-effort match against xll01 Personnel
      - tickets:      compliance dict from xll01 PersonnelCompetency — empty
                      {} when no match so consumers can still branch on
                      "did we have compliance data this run?" safely
      - enrichment fields: any additional fields (ty / sy / sum / newhire /
                      extras / drivers) from data/enrichment/<company>.json,
                      matched by name. Reserved keys (name/role/shift/mobile/
                      start/end/personnel_id/tickets) can't be clobbered.
    """
    try:
        import parse_macro_data as _pmd
    except Exception:                           # pragma: no cover — defensive
        _pmd = None

    enr_idx  = enrichment or {}
    reserved = {"name", "role", "shift", "mobile", "start", "end",
                "personnel_id", "tickets"}

    out: list[dict] = []
    for r in confirmed:
        entry: dict = {"name": r["name"], "role": r["role"]}
        if r.get("crew_type") and r["crew_type"] != "Unknown":
            entry["shift"] = r["crew_type"]
        if r.get("mobile"): entry["mobile"] = r["mobile"]
        if r.get("start"):  entry["start"]  = r["start"]
        if r.get("end"):    entry["end"]    = r["end"]
        if _pmd:
            pid = _pmd.match_personnel_id(r.get("first_name", ""),
                                          r.get("last_name", ""))
            if pid:
                entry["personnel_id"] = pid
                entry["tickets"]      = _pmd.tickets_for_person(
                    r.get("first_name", ""), r.get("last_name", ""))
            else:
                entry["tickets"]      = {}
        else:
            entry["tickets"] = {}
        enr = _enrichment_lookup(r.get("first_name", ""),
                                 r.get("last_name", ""),
                                 enr_idx)
        if enr:
            for k, v in enr.items():
                if k not in reserved:
                    entry[k] = v
        out.append(entry)
    return out


def build_shutdown(file_key: str, xlsx: pathlib.Path, rows: list[dict], fmt: str,
                   enrichment: dict[str, dict] | None = None) -> tuple[str, str, dict]:
    entry = ROSTER_MAP[file_key]
    company_key, client_name, project_label, site = entry[:4]
    shutdown_id_override = entry[4] if len(entry) > 4 else None
    confirmed = [r for r in rows if r["confirmed"]]
    if not confirmed:
        raise ValueError(f"{xlsx.name}: no confirmed rows")

    starts = [r["start"] for r in confirmed if r["start"]]
    ends   = [r["end"]   for r in confirmed if r["end"]]
    sd, ed = min(starts), max(ends)

    # Aggregates from the RosterCut file — retained as fallback and for the
    # auxiliary splits (crew/mobilised/labour-hire) that JobPlanningView
    # doesn't carry.
    rc_filled_by_role: dict[str, int] = {}
    crew_split:        dict[str, int] = {}
    mobilised_by_role: dict[str, int] = {}
    labour_hire_split: dict[str, int] = {}

    for r in confirmed:
        rc_filled_by_role[r["role"]]   = rc_filled_by_role.get(r["role"], 0) + 1
        crew_split[r["crew_type"]]     = crew_split.get(r["crew_type"], 0) + 1
        labour_hire_split[r["labour_hire"]] = labour_hire_split.get(r["labour_hire"], 0) + 1
        if r["mobilised"]:
            mobilised_by_role[r["role"]] = mobilised_by_role.get(r["role"], 0) + 1

    shutdown_id = shutdown_id_override or f"{company_key}-{sd[:7]}"

    # Rapid Crews is the source of truth for BOTH required and filled counts
    # when the JobNo is still in the macro workbook's JobPlanningView — this
    # keeps headline numbers in sync with the Rapid Crews website without
    # anyone having to re-export the RosterCut. The RosterCut XLSX is used
    # for the rich per-worker roster (Position-On-Project, Crew Type,
    # Confirmed flag) but its aggregate counts are a stale snapshot.
    planning_required: dict[str, int] | None = None
    planning_filled:   dict[str, int] | None = None
    if file_key.isdigit():
        try:
            import parse_macro_data as _pmd
            planning_required = _pmd.planning_required_for_jobno(int(file_key))
            planning_filled   = _pmd.planning_filled_for_jobno(int(file_key))
        except Exception as e:           # pragma: no cover — defensive
            print(f"  warn: macro planning lookup failed for {file_key}: {e}")
            planning_required = planning_filled = None

    if planning_required:
        # Both required + filled come from JobPlanningView. Keys that only
        # appear in the RosterCut roster are carried over at "0 required,
        # 0 filled" so the per-role table still surfaces them — RosterCut's
        # Position-On-Project names can be finer-grained than JobPlanningView
        # trades (e.g. "Fitter – Inspections" vs "Mechanical Fitter"), and
        # dropping the extras silently would hide real assignments.
        rc_only = set(rc_filled_by_role) - set(planning_required) - set(planning_filled or {})
        all_keys = set(planning_required) | set(planning_filled or {}) | rc_only
        required     = {r: int(planning_required.get(r, 0))           for r in all_keys}
        filled_final = {r: int((planning_filled or {}).get(r, 0))     for r in all_keys}
        target_source_meta = {"source": "rapid_crews_job_planning_view",
                              "job_no": int(file_key),
                              "total_required": sum(planning_required.values()),
                              "total_filled":   sum((planning_filled or {}).values())}
        required_target_source = "RAPID_CREWS_JOB_PLANNING"
    else:
        required, filled_final, target_source_meta = merge_targets(shutdown_id, rc_filled_by_role)
        target_exists = (TARGETS_DIR / f"{shutdown_id}.json").exists()
        required_target_source = "TARGET_FILE" if target_exists else "PLACEHOLDER_FROM_ROSTER"

    today       = dt.date.today()
    status      = _infer_status(dt.date.fromisoformat(sd),
                                dt.date.fromisoformat(ed), today)
    shutdown = {
        "id":               shutdown_id,
        "name":             project_label,
        "site":             site,
        "start_date":       sd,
        "end_date":         ed,
        "status":           status,
        "required_by_role": required,
        "filled_by_role":   filled_final,
        "crew_split":       crew_split,
        "mobilised_by_role": mobilised_by_role,
        "labour_hire_split": labour_hire_split,
        "roster":           _emit_roster_entries(confirmed, enrichment=enrichment),
        "_source": {
            "rapid_crews_roster_id":   file_key,
            "rapid_crews_export_file": xlsx.name,
            "source_format":           fmt,
            "required_target_source":  required_target_source,
            "rapid_crews_roster_size": len(confirmed),
            "rapid_crews_filled_by_role": rc_filled_by_role,   # RosterCut snapshot, preserved for audit
            "target_source":           target_source_meta,  # None for placeholder
        },
    }
    # Provenance for enriched rows — surfaced in data-quality warnings
    buckets = {k: 0 for k in ("explicit_column", "email_heuristic", "xref_exact", "unmatched")}
    ambiguous_samples: list[dict] = []
    for r in confirmed:
        nr = r.get("_name_resolution")
        if nr is None:
            continue
        if isinstance(nr, str) and nr.startswith("xref_ambiguous"):
            if len(ambiguous_samples) < 10:
                ambiguous_samples.append({"first_name": r["first_name"], "role": r["role"]})
            buckets["unmatched"] += 0     # counted separately below
            buckets.setdefault("xref_ambiguous", 0)
            buckets["xref_ambiguous"] += 1
        else:
            buckets[nr] = buckets.get(nr, 0) + 1
    if any(buckets.values()):
        shutdown["_source"]["name_resolution"] = {**buckets,
                                                  "ambiguous_samples": ambiguous_samples}
    return company_key, client_name, shutdown


# --------------------------------------------------------------------------- main

def _write_history_snapshots(triples: list[tuple[str, str, dict]]) -> None:
    """Persist one JSON snapshot per shutdown under data/history/<id>.json.

    These files are the safety net when Rapid Crews rolls a JobNo off the
    live SQL view: the next `parse_rapidcrews.py` run will see the snapshot,
    flag the shutdown as "archived", and re-hydrate it onto the dashboard.
    Files are overwritten only when the current run has something for that
    shutdown, so restored-from-archive shutdowns don't overwrite themselves
    with the same data every run.
    """
    HISTORY_DIR.mkdir(parents=True, exist_ok=True)
    now_iso = dt.datetime.now(dt.timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")
    for company_key, client_name, shutdown in triples:
        src = shutdown.get("_source", {})
        # Don't snapshot shutdowns that were themselves restored from archive
        # on this run — their data didn't freshen, so the file is already up
        # to date. `restored_from_archive` is stamped in _restore_from_history.
        if src.get("restored_from_archive"):
            continue
        snap = {
            "company_key":  company_key,
            "client_name":  client_name,
            "archived_at":  now_iso,
            "shutdown":     shutdown,
        }
        path = HISTORY_DIR / f"{shutdown['id']}.json"
        path.write_text(json.dumps(snap, indent=2))


def _restore_from_history(triples: list[tuple[str, str, dict]],
                          active_jobnos: set[int] | None
                          ) -> list[tuple[str, str, dict]]:
    """Fill in any ACTIVE_SHUTDOWNS JobNo that's missing from this run's output
    by reading the last-known snapshot from data/history/.

    Returns the list of restored triples (possibly empty). The caller is
    responsible for merging them into `combined`. Restored shutdowns get
    `status = "completed"` if their end_date is in the past, and their
    `_source.restored_from_archive = True` so the dashboard can flag them.
    """
    if not HISTORY_DIR.exists():
        return []
    present_jobnos: set[int] = set()
    for _, _, s in triples:
        src = s.get("_source", {})
        for key in ("rapid_crews_roster_id", "macro_data_job_no"):
            v = src.get(key)
            if v is None:
                continue
            try:
                present_jobnos.add(int(v))
            except (TypeError, ValueError):
                pass

    want = active_jobnos or set()
    missing = sorted(j for j in want if j not in present_jobnos)
    restored: list[tuple[str, str, dict]] = []
    today = dt.date.today()
    for job in missing:
        # Find the snapshot by JobNo — we don't know the shutdown_id upfront
        # since the file is named after the id, not the JobNo.
        snap_path = None
        for p in HISTORY_DIR.glob("*.json"):
            try:
                doc = json.loads(p.read_text())
            except (OSError, json.JSONDecodeError):
                continue
            src = doc.get("shutdown", {}).get("_source", {})
            for key in ("rapid_crews_roster_id", "macro_data_job_no"):
                v = src.get(key)
                try:
                    if v is not None and int(v) == job:
                        snap_path = p
                        break
                except (TypeError, ValueError):
                    pass
            if snap_path is not None:
                break
        if snap_path is None:
            print(f"  archive: JobNo {job} missing from this run and no snapshot "
                  f"in data/history/ — dashboard will show a gap")
            continue
        doc = json.loads(snap_path.read_text())
        shutdown = doc["shutdown"]
        # Re-infer status from the frozen dates, in case time has passed since
        # the snapshot was taken.
        try:
            sd = dt.date.fromisoformat(shutdown["start_date"])
            ed = dt.date.fromisoformat(shutdown["end_date"])
            shutdown["status"] = _infer_status(sd, ed, today)
        except (KeyError, ValueError):
            pass
        shutdown.setdefault("_source", {})
        shutdown["_source"]["restored_from_archive"] = True
        shutdown["_source"]["archive_path"]          = str(snap_path.relative_to(REPO_ROOT))
        shutdown["_source"]["archived_at"]           = doc.get("archived_at", "")
        restored.append((doc["company_key"], doc["client_name"], shutdown))
        print(f"  archive: JobNo {job:>4} -> restored from {snap_path.name} "
              f"({shutdown['id']} · roster={len(shutdown.get('roster', []))})")
    return restored


def _file_key(xlsx: pathlib.Path) -> str:
    """Lookup key into ROSTER_MAP.
      - Numeric leading token (Rapid Crews RosterCut) -> roster_id (e.g. "1353")
      - Anything else -> full filename stem, trimmed (e.g.
        "Tianqi Construction Ramp Up Project", "Kleenheat Major March 2026").
    Trailing whitespace in filenames is tolerated — the uploader for
    "Tianqi Construction Ramp Up Project .xlsx" accidentally left a space
    before the extension."""
    first = xlsx.name.split(" ", 1)[0]
    if first.isdigit():
        return first
    return xlsx.stem.strip()


def _merge_macro_triples(
    rc_triples: list[tuple[str, str, dict]],
    macro_triples: list[tuple[str, str, dict]],
) -> list[tuple[str, str, dict]]:
    """Merge macro-derived shutdowns into RosterCut triples.

    Rules:
      1) Macro overrides when JobNo matches an existing shutdown.
      2) If IDs collide but JobNo differs, macro shutdown id is suffixed
         with "-<job_no>" so both shutdowns survive.
      3) Preserve pointers to replaced RosterCut source files for audit.
    """
    combined: list[tuple[str, str, dict]] = list(rc_triples)

    def _job_no(sd: dict) -> int | None:
        src = sd.get("_source", {}) or {}
        for raw in (src.get("macro_data_job_no"), src.get("job_no"), src.get("rapid_crews_roster_id")):
            try:
                return int(raw)
            except (TypeError, ValueError):
                continue
        return None

    for company_key, client_name, shutdown in macro_triples:
        macro_job = _job_no(shutdown)
        sid = shutdown["id"]

        # Prefer authoritative identity (JobNo) over formatted id token.
        match_idx = None
        for i, (_, _, existing) in enumerate(combined):
            if macro_job is not None and _job_no(existing) == macro_job:
                match_idx = i
                break
        if match_idx is not None:
            _, _, rc_sd = combined[match_idx]
            print(f"  macro: JobNo {macro_job} -> {sid} overrides matched RosterCut")
            rc_src = rc_sd.get("_source", {}) or {}
            shutdown.setdefault("_source", {})["rapid_crews_export_file"] = rc_src.get("rapid_crews_export_file")
            shutdown["_source"]["rapid_crews_roster_id"] = rc_src.get("rapid_crews_roster_id")
            combined[match_idx] = (company_key, client_name, shutdown)
            continue

        # Different JobNo sharing the same monthly id (e.g. NAAN1 + NAAN2).
        if any(existing["id"] == sid for _, _, existing in combined) and macro_job is not None:
            shutdown = dict(shutdown)
            shutdown["id"] = f"{sid}-{macro_job}"
            print(f"  macro: JobNo {macro_job} id clash on {sid}; renamed -> {shutdown['id']}")
        else:
            print(f"  macro: JobNo {macro_job} -> {shutdown['id']} added (no matched RosterCut)")
        combined.append((company_key, client_name, shutdown))
    return combined


def _canonical_job_no(shutdown: dict) -> int | None:
    """Return the authoritative JobNo for filtering/merge decisions."""
    src = shutdown.get("_source", {}) or {}
    # For RosterCut-origin shutdowns, roster_id is authoritative.
    if src.get("source_format") == "rapidcrews":
        rid = src.get("rapid_crews_roster_id")
        try:
            return int(rid)
        except (TypeError, ValueError):
            pass
    for raw in (src.get("macro_data_job_no"), src.get("job_no"), src.get("rapid_crews_roster_id")):
        try:
            return int(raw)
        except (TypeError, ValueError):
            continue
    return None


def main() -> int:
    if not RAW_DIR.exists():
        print(f"No raw dir at {RAW_DIR}", file=sys.stderr)
        return 1

    # -- 1. Parse every mapped file first, so the Kleenheat enrichment step
    #       can cross-reference first-name + role against the other rosters.
    parsed: list[tuple[str, pathlib.Path, str, list[dict]]] = []
    rows_by_company: dict[str, list[dict]] = {}
    macro_path = RAW_DIR / "Rapidcrews Macro Data.xlsx"
    for xlsx in sorted(RAW_DIR.glob("*.xlsx")):
        # The SQL macro workbook lives in data/raw/ alongside RosterCut files;
        # parse_macro_data reads it separately, so it shouldn't be sniffed as
        # a RosterCut export (it isn't one, and the "unmapped" log is noise).
        if xlsx.resolve() == macro_path.resolve():
            continue
        key = _file_key(xlsx)
        if key not in ROSTER_MAP:
            print(f"  skip unmapped roster {key}: {xlsx.name}")
            continue
        fmt, rows = parse_roster(xlsx)
        company_key = ROSTER_MAP[key][0]
        parsed.append((key, xlsx, fmt, rows))
        rows_by_company.setdefault(company_key, []).extend(rows)

    if not parsed:
        print("No mapped roster files processed.", file=sys.stderr)
        return 1

    # -- 2. Enrich first-name-only rows (Kleenheat) with surnames from the
    #       RosterCut rosters, so retention can match across rosters.
    lookup = build_surname_lookup(rows_by_company)
    for key, xlsx, fmt, rows in parsed:
        if fmt != "kleenheat":
            continue
        stats = enrich_kleenheat_names(rows, lookup)
        print(f"  enrich {xlsx.name}: "
              f"{stats['explicit_column']} explicit col · "
              f"{stats['email_heuristic']} email · "
              f"{stats['xref_exact']} xref · "
              f"{stats['xref_ambiguous']} ambiguous · "
              f"{stats['unmatched']} unmatched")

    # -- 3. Load per-company enrichment overlays (resume prose, hand-curated
    #       annotations). Carries fields the Rapid Crews feed can't supply
    #       — trade years, shutdown years, resume narrative, newhire flags,
    #       driver's licence class. Per-site dashboards read these directly
    #       off the roster entries so they can stay thin renderers.
    enrichment_by_company: dict[str, dict[str, dict]] = {}
    for company_key in ("tronox", "covalent", "csbp"):
        enr = _load_enrichment(company_key)
        if enr:
            enrichment_by_company[company_key] = enr
            # index holds up to 3 keys per record — divide out for a
            # realistic record count when logging.
            approx = max(1, len(enr) // 3)
            print(f"  enrichment {company_key}: ~{approx} records "
                  f"({ENRICHMENT_DIR.relative_to(REPO_ROOT)}/{company_key}.json)")

    # -- 4. Build per-shutdown payloads from RosterCut files.
    rc_triples: list[tuple[str, str, dict]] = []
    for key, xlsx, fmt, rows in parsed:
        prelim_company_key = ROSTER_MAP[key][0]
        company_key, client_name, shutdown = build_shutdown(
            key, xlsx, rows, fmt,
            enrichment=enrichment_by_company.get(prelim_company_key))
        rc_triples.append((company_key, client_name, shutdown))
        print(f"  {key:>10}  {client_name:<10} {shutdown['id']:<22} "
              f"roster={len(shutdown['roster']):>3}  "
              f"{shutdown['start_date']} → {shutdown['end_date']}  "
              f"[{shutdown['status']}]")

    # -- 3b. Pull shutdowns from Rapidcrews Macro Data.xlsx (ACTIVE_SHUTDOWNS
    #        sheet), then merge with explicit JobNo-first conflict rules.
    import parse_macro_data
    macro_triples   = parse_macro_data.shutdowns_from_macro_data()
    active_jobnos   = parse_macro_data.active_shutdowns_jobnos()
    combined = _merge_macro_triples(rc_triples, macro_triples)
    # Historical safety: old snapshots sometimes stamped a macro JobNo onto a
    # RosterCut shutdown with a different roster_id. Keep source fields sane.
    for _, _, shutdown in combined:
        src = shutdown.get("_source", {}) or {}
        if src.get("source_format") != "rapidcrews":
            continue
        try:
            rid = int(src.get("rapid_crews_roster_id"))
        except (TypeError, ValueError):
            continue
        bad_macro = src.get("macro_data_job_no")
        try:
            if bad_macro is not None and int(bad_macro) != rid:
                src.pop("macro_data_job_no", None)
        except (TypeError, ValueError):
            src.pop("macro_data_job_no", None)

    # -- 3c. If the ACTIVE_SHUTDOWNS sheet is present, it's an allow-list:
    #        RosterCut shutdowns whose numeric roster_id isn't listed drop
    #        out. Non-numeric rosters (Kleenheat / Pegasus historicals)
    #        always pass — they're retention seeds, not JobNo-driven.
    if active_jobnos is not None:
        kept: list[tuple[str, str, dict]] = []
        for company_key, client_name, shutdown in combined:
            job_no = _canonical_job_no(shutdown)
            if job_no is None or job_no in active_jobnos:
                kept.append((company_key, client_name, shutdown))
            else:
                print(f"  filter: {shutdown['id']} (JobNo {job_no}) "
                      f"not in ACTIVE_SHUTDOWNS — dropping")
        combined = kept

    # -- 3d. Historical retention: for any ACTIVE_SHUTDOWNS JobNo that produced
    #        no shutdown this run AND has no matching raw XLSX, try restoring
    #        the last-known snapshot from data/history/. This keeps the tile
    #        on the dashboard after Rapid Crews' live SQL view rolls the
    #        JobNo off (it's time-windowed to roughly the next 12 months), so
    #        the ops team doesn't suddenly lose visibility on a shutdown
    #        that's still operationally relevant.
    restored = _restore_from_history(combined, active_jobnos)
    for triple in restored:
        combined.append(triple)

    by_company: dict[str, dict] = {}
    for company_key, client_name, shutdown in combined:
        by_company.setdefault(company_key, {"company": client_name, "shutdowns": []})
        by_company[company_key]["shutdowns"].append(shutdown)

    # -- 3e. Snapshot every shutdown we just built into data/history/. Files
    #        are committed alongside data/*.json so they're safe across
    #        branches and in the repo's audit trail.
    _write_history_snapshots(combined)

    # -- 4. Write per-company JSON files.
    now = dt.datetime.now(dt.timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")
    for company_key, payload in by_company.items():
        payload["generated_at"] = now
        payload["shutdowns"].sort(key=lambda s: s["start_date"])
        out = DATA_DIR / f"{company_key}.json"
        out.write_text(json.dumps(payload, indent=2))
        total = sum(len(s["roster"]) for s in payload["shutdowns"])
        print(f"Wrote {out.relative_to(REPO_ROOT)}: "
              f"{len(payload['shutdowns'])} shutdown(s), {total} confirmed heads")

    # -- 4b. Write consolidated resumes JSON. Kept separate from the per-
    #        company payloads so the dashboard can fetch it once and
    #        decorate workers in every tab (matrix, ops roster).
    try:
        import parse_macro_data as _pmd
        resumes = _pmd.resumes_from_macro_data()
    except Exception as e:
        print(f"  warn: resumes lookup failed: {e}")
        resumes = []
    resumes_doc = {
        "generated_at": now,
        "source_file":  "Resumes.xlsx",
        "resumes":      resumes,
    }
    (DATA_DIR / "resumes.json").write_text(json.dumps(resumes_doc, indent=2))
    print(f"Wrote data/resumes.json: {len(resumes)} resume link(s)")

    # -- 5. Backfill empty payloads for any client the dashboard lists but
    #       which got no rosters this run (prevents 404s on page load).
    referenced = {"covalent", "tronox", "csbp"}
    for company_key in referenced - by_company.keys():
        path = DATA_DIR / f"{company_key}.json"
        if not path.exists():
            payload = {"company": company_key.title(), "generated_at": now, "shutdowns": []}
            path.write_text(json.dumps(payload, indent=2))
            print(f"Wrote {path.relative_to(REPO_ROOT)}: empty (no roster supplied)")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
