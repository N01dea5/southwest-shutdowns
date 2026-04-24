#!/usr/bin/env python3
"""Normalise the RapidCrews workbook before parsing.

The dashboard parser historically consumed `xpbi02 PersonnelRosterView`.
The newer RapidCrews report emits roster rows in `xpbi02 DailyPersonnelSchedule`
and resolves clients via `xpbi02 ClientView`.

This script creates a backwards-compatible `xpbi02 PersonnelRosterView` sheet
from the new daily schedule, applying the new source rules before the existing
parser runs:

- resolve Client GUIDs through `xpbi02 ClientView`
- de-duplicate to one row per (Job No, Personnel Id, Schedule Date), preferring
  onsite over demobilised when both are present
- skip OnSite = 0 rows only when the row date is before today
- for today/future rows, pass through any non-rejected row regardless of OnSite
- when a worker is rostered every calendar day of a month, keep 50% of those
  daily rows to avoid overcounting continuous-contract workers

If the workbook still has the old `xpbi02 PersonnelRosterView`, the script
leaves it alone except for numbered-sheet fallback handling.
"""
from __future__ import annotations

import calendar
import datetime as dt
import pathlib
import re
from collections import defaultdict
from typing import Any

import openpyxl

REPO_ROOT = pathlib.Path(__file__).resolve().parent.parent
MACRO_FILE = REPO_ROOT / "data" / "raw" / "Rapidcrews Macro Data.xlsx"

DAILY_SCHEDULE_SHEET = "xpbi02 DailyPersonnelSchedule"
ROSTER_VIEW_SHEET = "xpbi02 PersonnelRosterView"
CLIENT_VIEW_SHEET = "xpbi02 ClientView"

COMPAT_HEADERS = [
    "Job No",
    "Client",
    "Site",
    "Personnel Id",
    "Schedule Date",
    "Schedule Type",
    "IsOnLocation",
]

ALIASES = {
    "Job No": ["Job No", "JobNo", "Job", "Job Number", "JobNo_"],
    "Client": ["Client", "ClientId", "Client Id", "ClientID"],
    "Site": ["Site", "SiteName", "Site Name", "Location", "Job Site"],
    "Personnel Id": ["Personnel Id", "PersonnelId", "Personnel ID", "Employee Id", "EmployeeID", "Resource Id"],
    "Schedule Date": ["Schedule Date", "Date", "Roster Date", "RosterDate", "Shift Date", "Work Date"],
    "Schedule Type": ["Schedule Type", "ScheduleType", "Shift", "Shift Type", "Roster Type", "Crew", "Crew Type"],
    "Status": ["Status", "Roster Status", "Schedule Status", "Personnel Status", "Booking Status"],
    "OnSite": ["OnSite", "On Site", "IsOnLocation", "Is On Location", "On Location"],
    "ClientId": ["ClientId", "Client Id", "ClientID"],
    "ClientName": ["ClientName", "Client Name", "Name"],
}

REJECTED_TERMS = ("reject", "declin", "cancel", "remove")
DEMOB_TERMS = ("demob", "offsite", "off site")
ONSITE_TERMS = ("onsite", "on site", "mobilis", "confirmed", "booked", "working")


def _clean(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\xa0", " ")).strip()


def _norm(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", _clean(value).lower())


def _resolve_sheet(wb: openpyxl.Workbook, canonical: str) -> str | None:
    if canonical in wb.sheetnames:
        return canonical
    target = _norm(canonical)
    for name in wb.sheetnames:
        n = _norm(name)
        if n == target or re.fullmatch(rf"{re.escape(target)}\d+", n):
            return name
    return None


def _header_index(ws) -> dict[str, int]:
    headers = [c.value for c in next(ws.iter_rows(max_row=1))]
    return {_clean(h): i for i, h in enumerate(headers) if _clean(h)}


def _find_col(idx: dict[str, int], canonical: str, required: bool = True) -> int | None:
    normal = {_norm(k): v for k, v in idx.items()}
    for alias in ALIASES.get(canonical, [canonical]):
        if _norm(alias) in normal:
            return normal[_norm(alias)]
    if required:
        raise KeyError(f"Missing required column for {canonical!r}; available: {', '.join(idx)}")
    return None


def _get(row: tuple[Any, ...], idx: int | None) -> Any:
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _parse_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    text = _clean(value).lower()
    return text in {"1", "true", "yes", "y", "on", "onsite", "on site"}


def _parse_date(value: Any) -> dt.date | None:
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    text = _clean(value)
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return dt.datetime.strptime(text[:10], fmt).date()
        except ValueError:
            pass
    return None


def _parse_job(value: Any) -> int | None:
    if isinstance(value, (int, float)):
        return int(value)
    m = re.search(r"\d+", _clean(value))
    return int(m.group(0)) if m else None


def _load_client_lookup(wb: openpyxl.Workbook) -> dict[str, str]:
    sheet = _resolve_sheet(wb, CLIENT_VIEW_SHEET)
    if not sheet:
        return {}
    ws = wb[sheet]
    idx = _header_index(ws)
    id_col = _find_col(idx, "ClientId")
    name_col = _find_col(idx, "ClientName")
    out: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        cid = _clean(_get(row, id_col))
        name = _clean(_get(row, name_col))
        if cid and name:
            out[cid] = name
            out[_norm(cid)] = name
    return out


def _resolve_client(raw: Any, client_lookup: dict[str, str]) -> str:
    text = _clean(raw)
    return client_lookup.get(text) or client_lookup.get(_norm(text)) or text


def _compat_client_site(client_name: str, site_raw: Any) -> tuple[str, str]:
    site = _clean(site_raw)
    c = client_name.lower()
    s = site.lower()
    if "tronox" in c or "tronox" in s:
        return "SOUTH WEST", "Tronox"
    if "covalent" in c or "covalent" in s or "mt holland" in c or "mt holland" in s:
        return "SOUTH WEST", "Covalent Lithium"
    if "kleenheat" in c or "kleenheat" in s or "kpf" in c or "kpf" in s:
        return "SOUTH WEST", "Kleenheat"
    if "csbp" in c or "csbp" in s:
        return "CSBP", "CSBP Kwinana"
    return client_name or "SOUTH WEST", site or client_name


def _status_text(row: tuple[Any, ...], status_col: int | None) -> str:
    return _clean(_get(row, status_col)).lower()


def _is_rejected(status: str) -> bool:
    return any(term in status for term in REJECTED_TERMS)


def _rank(row: tuple[Any, ...], status_col: int | None, onsite_col: int | None) -> int:
    status = _status_text(row, status_col)
    onsite = _parse_bool(_get(row, onsite_col))
    if _is_rejected(status):
        return -1
    if onsite or "onsite" in status or "on site" in status:
        return 40
    if any(term in status for term in ONSITE_TERMS):
        return 30
    if any(term in status for term in DEMOB_TERMS):
        return 10
    return 20


def _schedule_type(raw: Any, status: str) -> str:
    text = _clean(raw)
    lower = text.lower()
    if "night" in lower:
        return "Night Shift"
    if "day" in lower:
        return "Day Shift"
    if lower == "rnr" or "r&r" in lower:
        return "RNR"
    if "night" in status:
        return "Night Shift"
    if "rnr" in status or "r&r" in status:
        return "RNR"
    return "Day Shift"


def _should_include(row_date: dt.date, status: str, onsite: bool, today: dt.date) -> bool:
    if _is_rejected(status):
        return False
    if row_date < today and not onsite:
        return False
    return True


def _scale_full_month_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[tuple[int, str, int, int], list[dict[str, Any]]] = defaultdict(list)
    for item in rows:
        d: dt.date = item["date"]
        grouped[(item["job"], item["pid"], d.year, d.month)].append(item)

    keep_ids: set[int] = set()
    for (_job, _pid, year, month), items in grouped.items():
        by_date = {item["date"]: item for item in items}
        days_in_month = calendar.monthrange(year, month)[1]
        all_month_dates = {dt.date(year, month, day) for day in range(1, days_in_month + 1)}
        ordered = [by_date[d] for d in sorted(by_date)]
        if set(by_date) == all_month_dates:
            # 50% actual attendance assumption for continuous-contract workers.
            for item in ordered[::2]:
                keep_ids.add(id(item))
        else:
            for item in ordered:
                keep_ids.add(id(item))
    return [item for item in rows if id(item) in keep_ids]


def _write_compat_roster(wb: openpyxl.Workbook, rows: list[dict[str, Any]]) -> None:
    if ROSTER_VIEW_SHEET in wb.sheetnames:
        del wb[ROSTER_VIEW_SHEET]
    ws = wb.create_sheet(ROSTER_VIEW_SHEET)
    ws.append(COMPAT_HEADERS)
    for item in rows:
        ws.append([
            item["job"],
            item["client"],
            item["site"],
            item["pid"],
            item["date"],
            item["schedule_type"],
            item["onsite"],
        ])


def _normalise_numbered_sheet_names(wb: openpyxl.Workbook) -> bool:
    changed = False
    canonical_names = [
        "ACTIVE_SHUTDOWNS",
        "xpbi02 JobPlanningView",
        "xpbi02 DisciplineTrade",
        "xll01 Personnel",
        "xll01 PersonnelCompetency",
        "xpbi02 PersonnelCalendarView",
        CLIENT_VIEW_SHEET,
    ]
    for canonical in canonical_names:
        if canonical in wb.sheetnames:
            continue
        found = _resolve_sheet(wb, canonical)
        if found and found != canonical:
            wb[found].title = canonical
            print(f"normalise_rapidcrews_workbook: renamed {found!r} -> {canonical!r}")
            changed = True
    return changed


def _normalise_daily_schedule(wb: openpyxl.Workbook) -> bool:
    daily_sheet = _resolve_sheet(wb, DAILY_SCHEDULE_SHEET)
    if not daily_sheet:
        # Backwards compatibility: if only a numbered old sheet exists, rename it.
        old_sheet = _resolve_sheet(wb, ROSTER_VIEW_SHEET)
        if old_sheet and old_sheet != ROSTER_VIEW_SHEET and ROSTER_VIEW_SHEET not in wb.sheetnames:
            wb[old_sheet].title = ROSTER_VIEW_SHEET
            print(f"normalise_rapidcrews_workbook: renamed {old_sheet!r} -> {ROSTER_VIEW_SHEET!r}")
            return True
        print("normalise_rapidcrews_workbook: no DailyPersonnelSchedule sheet found; using existing roster view")
        return False

    ws = wb[daily_sheet]
    idx = _header_index(ws)
    job_col = _find_col(idx, "Job No")
    client_col = _find_col(idx, "Client")
    site_col = _find_col(idx, "Site", required=False)
    pid_col = _find_col(idx, "Personnel Id")
    date_col = _find_col(idx, "Schedule Date")
    sched_col = _find_col(idx, "Schedule Type", required=False)
    status_col = _find_col(idx, "Status", required=False)
    onsite_col = _find_col(idx, "OnSite", required=False)

    client_lookup = _load_client_lookup(wb)
    today = dt.date.today()
    dedup: dict[tuple[int, str, dt.date], tuple[int, tuple[Any, ...]]] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not any(row):
            continue
        job = _parse_job(_get(row, job_col))
        pid = _clean(_get(row, pid_col))
        row_date = _parse_date(_get(row, date_col))
        if not job or not pid or not row_date:
            continue
        status = _status_text(row, status_col)
        onsite = _parse_bool(_get(row, onsite_col))
        if not _should_include(row_date, status, onsite, today):
            continue
        key = (job, pid, row_date)
        rank = _rank(row, status_col, onsite_col)
        if key not in dedup or rank > dedup[key][0]:
            dedup[key] = (rank, row)

    compat_rows: list[dict[str, Any]] = []
    for (job, pid, row_date), (_rank_value, row) in sorted(dedup.items(), key=lambda x: (x[0][0], x[0][2], x[0][1])):
        status = _status_text(row, status_col)
        client_name = _resolve_client(_get(row, client_col), client_lookup)
        client, site = _compat_client_site(client_name, _get(row, site_col))
        compat_rows.append({
            "job": job,
            "client": client,
            "site": site,
            "pid": pid,
            "date": row_date,
            "schedule_type": _schedule_type(_get(row, sched_col), status),
            "onsite": _parse_bool(_get(row, onsite_col)),
        })

    compat_rows = _scale_full_month_rows(compat_rows)
    _write_compat_roster(wb, compat_rows)
    print(f"normalise_rapidcrews_workbook: wrote {len(compat_rows)} rows to {ROSTER_VIEW_SHEET!r} from {daily_sheet!r}")
    return True


def main() -> int:
    if not MACRO_FILE.exists():
        print("normalise_rapidcrews_workbook: macro workbook not found; skipped")
        return 0

    wb = openpyxl.load_workbook(MACRO_FILE)
    try:
        changed = _normalise_numbered_sheet_names(wb)
        changed = _normalise_daily_schedule(wb) or changed
        if changed:
            wb.save(MACRO_FILE)
            print(f"normalise_rapidcrews_workbook: saved {MACRO_FILE.relative_to(REPO_ROOT)}")
        else:
            print("normalise_rapidcrews_workbook: no workbook changes required")
    finally:
        wb.close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
