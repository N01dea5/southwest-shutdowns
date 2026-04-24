#!/usr/bin/env python3
from __future__ import annotations

import json
import pathlib
import re
from typing import Any

import openpyxl

REPO_ROOT = pathlib.Path(__file__).resolve().parent.parent
RAW_DIR = REPO_ROOT / "data" / "raw"
AUDIT_DIR = REPO_ROOT / "data" / "audit"
OUT_JSON = AUDIT_DIR / "rapidcrews_workbook_schema.json"
OUT_MD = AUDIT_DIR / "rapidcrews_workbook_schema.md"

EXPECTED = {
    "ACTIVE_SHUTDOWNS": ["JobNo"],
    "xpbi02 JobPlanningView": ["JobNo", "CompetencyId", "Required", "Filled"],
    "xpbi02 PersonnelRosterView": ["Job No", "Client", "Site", "Personnel Id", "Schedule Date", "Schedule Type", "IsOnLocation"],
    "xpbi02 DisciplineTrade": ["TradeId", "Trade"],
    "xll01 Personnel": ["Personnel Id", "Given Names", "Surname", "Primary Role", "Mobile", "Hire Company"],
    "xll01 PersonnelCompetency": ["Personnel Id", "Competency", "Expiry", "Document Location", "Archived"],
    "xpbi02 PersonnelCalendarView": ["Personnel Id", "Start Date", "End Date"],
}

ALIASES = {
    "job no": "Job No",
    "jobno": "JobNo",
    "personnelid": "Personnel Id",
    "personnel id": "Personnel Id",
    "employee id": "Personnel Id",
    "given names": "Given Names",
    "first name": "Given Names",
    "surname": "Surname",
    "last name": "Surname",
    "primary role": "Primary Role",
    "role": "Primary Role",
    "hire company": "Hire Company",
    "hiring company": "Hire Company",
    "schedule date": "Schedule Date",
    "date": "Schedule Date",
    "schedule type": "Schedule Type",
    "isonlocation": "IsOnLocation",
    "is on location": "IsOnLocation",
    "competencyid": "CompetencyId",
    "tradeid": "TradeId",
    "trade": "Trade",
    "competency": "Competency",
    "document location": "Document Location",
    "start date": "Start Date",
    "end date": "End Date",
}

KEYWORDS = set(k.lower() for vals in EXPECTED.values() for k in vals) | set(ALIASES)


def clean(v: Any) -> str:
    return re.sub(r"\s+", " ", str(v or "").replace("\xa0", " ")).strip()


def norm(v: Any) -> str:
    return re.sub(r"[^a-z0-9]+", " ", clean(v).lower()).strip()


def canon(v: Any) -> str:
    n = norm(v)
    return ALIASES.get(n, clean(v))


def row_values(ws, row_no: int, max_cols: int = 80) -> list[str]:
    return [clean(c.value) for c in ws[row_no][:max_cols]]


def score_header(row: list[str]) -> int:
    vals = [norm(x) for x in row if clean(x)]
    keyword_hits = sum(1 for x in vals if x in KEYWORDS or ALIASES.get(x))
    return keyword_hits * 10 + len(vals)


def find_header_row(ws) -> tuple[int, list[str]]:
    best = (1, [], -1)
    for r in range(1, min(ws.max_row, 20) + 1):
        vals = row_values(ws, r)
        score = score_header(vals)
        if score > best[2]:
            best = (r, vals, score)
    return best[0], [v for v in best[1] if v]


def missing_headers(headers: list[str], expected: list[str]) -> list[str]:
    have = {canon(h) for h in headers}
    return [h for h in expected if h not in have]


def inspect_workbook(path: pathlib.Path) -> dict[str, Any]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        sheets = []
        for ws in wb.worksheets:
            header_row, headers = find_header_row(ws)
            expected = EXPECTED.get(ws.title, [])
            sheets.append({
                "sheet": ws.title,
                "rows": ws.max_row,
                "columns": ws.max_column,
                "detected_header_row": header_row,
                "headers": headers,
                "expected_headers": expected,
                "missing_expected_headers": missing_headers(headers, expected) if expected else [],
                "known_sheet": ws.title in EXPECTED,
            })
        return {"file": str(path.relative_to(REPO_ROOT)), "sheets": sheets}
    finally:
        wb.close()


def write_markdown(payload: dict[str, Any]) -> None:
    lines = ["# RapidCrews workbook schema audit", ""]
    for wb in payload["workbooks"]:
        lines += [f"## `{wb['file']}`", ""]
        for s in wb["sheets"]:
            status = "OK" if s["known_sheet"] and not s["missing_expected_headers"] else ("NEW" if not s["known_sheet"] else "CHECK")
            lines.append(f"### {status} — `{s['sheet']}`")
            lines.append(f"Rows: {s['rows']} | Columns: {s['columns']} | Header row: {s['detected_header_row']}")
            if s["missing_expected_headers"]:
                lines.append(f"Missing expected headers: {', '.join(s['missing_expected_headers'])}")
            lines.append("Headers: " + ", ".join(f"`{h}`" for h in s["headers"][:80]))
            lines.append("")
    OUT_MD.write_text("\n".join(lines))


def main() -> int:
    AUDIT_DIR.mkdir(parents=True, exist_ok=True)
    files = sorted(RAW_DIR.glob("*.xlsx")) + sorted(RAW_DIR.glob("*.xlsm"))
    payload = {"workbooks": []}
    for path in files:
        try:
            payload["workbooks"].append(inspect_workbook(path))
        except Exception as exc:
            payload["workbooks"].append({"file": str(path.relative_to(REPO_ROOT)), "error": str(exc), "sheets": []})
    OUT_JSON.write_text(json.dumps(payload, indent=2))
    write_markdown(payload)
    print(f"audit_rapidcrews_workbook: wrote {OUT_JSON.relative_to(REPO_ROOT)} and {OUT_MD.relative_to(REPO_ROOT)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
