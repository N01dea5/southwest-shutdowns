#!/usr/bin/env python3
"""Append rejected/declined RapidCrews shutdown rows to personnel_calendar.json.

This feeds the existing internal worker matrix availability overlay. It does
not create a new dashboard table.
"""
from __future__ import annotations

import datetime as dt
import json
import pathlib
import re
from typing import Any

import openpyxl

REPO_ROOT = pathlib.Path(__file__).resolve().parent.parent
DATA_DIR = REPO_ROOT / "data"
RAW_FILE = DATA_DIR / "raw" / "Rapidcrews Macro Data.xlsx"
OUT_FILE = DATA_DIR / "personnel_calendar.json"
DAILY_SHEET = "xpbi02 DailyPersonnelSchedule"
PERSONNEL_SHEET = "xll01 Personnel"
CLIENT_SHEET = "xpbi02 ClientView"

REJECT_TERMS = ("reject", "rejected", "declin", "declined", "turn down", "turned down")

ALIASES = {
    "Personnel Id": ["Personnel Id", "PersonnelId", "Personnel ID", "Employee Id", "EmployeeID", "ResourceId", "Resource Id"],
    "First Name": ["FirstName", "First Name", "Given Names", "Given Name"],
    "Surname": ["Surname", "Last Name", "Family Name"],
    "Status": ["Status", "Roster Status", "Schedule Status", "Booking Status", "jobStatus", "Job Status"],
    "Date": ["ReportDate", "Report Date", "Schedule Date", "Date", "Roster Date"],
    "Job No": ["QuoteNo", "Quote No", "OrderNo", "Order No", "Job No", "JobNo", "JobId", "Job Id"],
    "Client": ["Client", "ClientId", "Client Id", "ClientID"],
    "Site": ["Site", "SiteId", "Site Id", "SiteName", "Site Name"],
    "Role": ["Trade", "Discipline", "Primary Role", "Role", "Position"],
    "ClientId": ["ClientId", "Client Id", "ClientID", "Id"],
    "ClientName": ["ClientName", "Client Name", "Name"],
}


def _clean(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\xa0", " ")).strip()


def _norm(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", _clean(value).lower())


def _name_key(*parts: Any) -> str:
    return re.sub(r"[^a-z]+", "", " ".join(_clean(p) for p in parts if _clean(p)).lower())


def _resolve_sheet(wb: openpyxl.Workbook, canonical: str) -> str | None:
    if canonical in wb.sheetnames:
        return canonical
    target = _norm(canonical)
    for name in wb.sheetnames:
        n = _norm(name)
        if n == target or re.fullmatch(rf"{re.escape(target)}\d+", n):
            return name
    return None


def _headers(ws) -> dict[str, int]:
    try:
        header = list(next(ws.iter_rows(max_row=1, values_only=True)))
    except StopIteration:
        return {}
    return {_clean(h): i for i, h in enumerate(header) if _clean(h)}


def _find_col(headers: dict[str, int], canonical: str, required: bool = True) -> int | None:
    normal = {_norm(k): v for k, v in headers.items()}
    for alias in ALIASES.get(canonical, [canonical]):
        if _norm(alias) in normal:
            return normal[_norm(alias)]
    if required:
        raise KeyError(f"Missing required column {canonical!r}; available: {', '.join(headers)}")
    return None


def _get(row: tuple[Any, ...], idx: int | None) -> Any:
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _date(value: Any) -> str:
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    text = _clean(value)
    if not text:
        return ""
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return dt.datetime.strptime(text[:10], fmt).date().isoformat()
        except ValueError:
            pass
    return ""


def _job_no(value: Any) -> str:
    text = _clean(value)
    matches = re.findall(r"\b\d{3,6}\b", text)
    return matches[0] if matches else text


def _is_rejected(status: str) -> bool:
    lowered = _clean(status).lower()
    return any(term in lowered for term in REJECT_TERMS)


def _load_personnel(wb: openpyxl.Workbook) -> dict[str, dict[str, str]]:
    sheet = _resolve_sheet(wb, PERSONNEL_SHEET)
    if not sheet:
        return {}
    ws = wb[sheet]
    headers = _headers(ws)
    pid_i = _find_col(headers, "Personnel Id", required=False)
    first_i = _find_col(headers, "First Name", required=False)
    last_i = _find_col(headers, "Surname", required=False)
    role_i = _find_col(headers, "Role", required=False)
    out: dict[str, dict[str, str]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        pid = _clean(_get(row, pid_i))
        if not pid:
            continue
        first = _clean(_get(row, first_i))
        last = _clean(_get(row, last_i))
        name = _clean(f"{first} {last}")
        out[pid] = {"name": name, "role": _clean(_get(row, role_i))}
    return out


def _load_clients(wb: openpyxl.Workbook) -> dict[str, str]:
    sheet = _resolve_sheet(wb, CLIENT_SHEET)
    if not sheet:
        return {}
    ws = wb[sheet]
    headers = _headers(ws)
    id_i = _find_col(headers, "ClientId", required=False)
    name_i = _find_col(headers, "ClientName", required=False)
    out: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        cid = _clean(_get(row, id_i))
        name = _clean(_get(row, name_i))
        if cid and name:
            out[cid] = name
            out[_norm(cid)] = name
    return out


def _client(raw: Any, clients: dict[str, str]) -> str:
    text = _clean(raw)
    return clients.get(text) or clients.get(_norm(text)) or text


def _read_rejections() -> list[dict[str, Any]]:
    if not RAW_FILE.exists():
        print("apply_rejected_shutdowns: workbook not found; skipped")
        return []
    wb = openpyxl.load_workbook(RAW_FILE, data_only=True, read_only=True)
    try:
        sheet = _resolve_sheet(wb, DAILY_SHEET)
        if not sheet:
            print("apply_rejected_shutdowns: DailyPersonnelSchedule not found; skipped")
            return []
        personnel = _load_personnel(wb)
        clients = _load_clients(wb)
        ws = wb[sheet]
        headers = _headers(ws)
        pid_i = _find_col(headers, "Personnel Id")
        first_i = _find_col(headers, "First Name", required=False)
        last_i = _find_col(headers, "Surname", required=False)
        status_i = _find_col(headers, "Status", required=False)
        date_i = _find_col(headers, "Date")
        job_i = _find_col(headers, "Job No", required=False)
        client_i = _find_col(headers, "Client", required=False)
        site_i = _find_col(headers, "Site", required=False)
        role_i = _find_col(headers, "Role", required=False)

        if status_i is None:
            print("apply_rejected_shutdowns: no status column found; no rejected rows added")
            return []

        seen: set[tuple[str, str, str]] = set()
        events: list[dict[str, Any]] = []
        for row_no, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            status = _clean(_get(row, status_i))
            if not _is_rejected(status):
                continue
            pid = _clean(_get(row, pid_i))
            day = _date(_get(row, date_i))
            if not pid or not day:
                continue
            job = _job_no(_get(row, job_i))
            key = (pid, day, job)
            if key in seen:
                continue
            seen.add(key)
            first = _clean(_get(row, first_i))
            last = _clean(_get(row, last_i))
            person = personnel.get(pid, {})
            name = _clean(f"{first} {last}") or person.get("name", "")
            if not name:
                continue
            company = _client(_get(row, client_i), clients)
            site = _clean(_get(row, site_i))
            role = _clean(_get(row, role_i)) or person.get("role", "")
            desc_bits = ["Rejected/declined shutdown", company, site, f"Job {job}" if job else "", status]
            events.append({
                "personnel_id": pid,
                "name": name,
                "name_key": _name_key(name),
                "role": role,
                "start": day,
                "end": day,
                "type": "rejected_shutdown",
                "description": " · ".join(b for b in desc_bits if b),
                "company_or_job": " · ".join(b for b in (company, site) if b),
                "job_no": job,
                "source_row": row_no,
            })
        return events
    finally:
        wb.close()


def main() -> int:
    payload = json.loads(OUT_FILE.read_text()) if OUT_FILE.exists() else {"events": []}
    existing = payload.setdefault("events", [])
    existing_keys = {(e.get("personnel_id"), e.get("start"), e.get("job_no"), e.get("type")) for e in existing}
    added = 0
    for event in _read_rejections():
        key = (event.get("personnel_id"), event.get("start"), event.get("job_no"), event.get("type"))
        if key in existing_keys:
            continue
        existing.append(event)
        existing_keys.add(key)
        added += 1
    payload["generated_at"] = dt.datetime.utcnow().replace(microsecond=0).isoformat() + "Z"
    payload["includes_rejected_shutdowns"] = True
    OUT_FILE.write_text(json.dumps(payload, indent=2))
    print(f"apply_rejected_shutdowns: added {added} rejected shutdown events")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
