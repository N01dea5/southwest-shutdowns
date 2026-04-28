#!/usr/bin/env python3
"""Build dashboard shutdowns from `Rapidcrews Macro Data.xlsx`.

End-user workflow
-----------------
Open `Rapidcrews Macro Data.xlsx`, go to the `ACTIVE_SHUTDOWNS` sheet, add or
delete rows in the `JobNo` column. Each row = one shutdown the dashboard
should show. When the file lands in the repo the GitHub Action reparses and
the dashboard updates — no code changes, no RosterCut drops required.

How the sheet interacts with the existing RosterCut flow
--------------------------------------------------------
- Sheet missing or no data rows -> legacy behaviour: every RosterCut file in
  `data/raw/` still appears exactly as before.
- Sheet present with rows -> acts as an allow-list:
    * RosterCut shutdowns whose numeric roster_id is in the list pass through
      (they're richer than macro-derived data — Position-On-Project, Crew
      Type, Confirmed flag — so they win on collisions).
    * JobNos in the list without a matching RosterCut file get built from the
      macro workbook's JobPlanningView + PersonnelRosterView.
    * Non-numeric rosters (Kleenheat historical / Pegasus imports) always
      pass — they're not keyed by JobNo.

Macro-derived shutdowns have slightly lower fidelity than RosterCut ones —
per-worker `role` is the employee's Primary Role (from `xll01 Personnel`)
rather than Position-On-Project, and there's no "Contingency" crew bucket.
JobPlanningView still drives the Required/Filled aggregates accurately.
"""
from __future__ import annotations

import datetime as dt
import pathlib
from collections import Counter, defaultdict

import openpyxl

import parse_rapidcrews as rc


REPO_ROOT    = pathlib.Path(__file__).resolve().parent.parent
MACRO_FILE   = REPO_ROOT / "data" / "raw" / "Rapidcrews Macro Data.xlsx"
RESUMES_FILE = REPO_ROOT / "Resumes.xlsx"   # standalone so end users can own
                                            # it on SharePoint without touching
                                            # the Rapid Crews SQL export

CONTROL_SHEET      = "ACTIVE_SHUTDOWNS"
JOB_PLANNING_SHEET = "xpbi02 JobPlanningView"
ROSTER_VIEW_SHEET  = "xpbi02 PersonnelRosterView"
DAILY_SCHEDULE_SHEET = "xpbi02 DailyPersonnelSchedule"
TRADE_SHEET        = "xpbi02 DisciplineTrade"
PERSONNEL_SHEET    = "xll01 Personnel"
COMPLIANCE_SHEET   = "xll01 PersonnelCompetency"

# DailyPersonnelSchedule statuses that mean "this person is filling a role on
# this job". Anything else (Contacted, Short List, Declined, Late Withdrawal,
# Rejected, Planning) is upstream-of-confirmation noise and shouldn't appear
# in the named roster — that's how the matrix and shutdown-detail "filled"
# stay aligned with RapidCrews JobPlanningView's Filled count.
ONSITE_PERSONNEL_STATUSES = {"Confirmed", "Mobilising", "Onsite", "Demobilised"}

# The SQL DisciplineTrade table uses "Rigger - Advanced/Intermediate/Basic"
# but the RosterCut Position-On-Project (and every downstream key — target
# files, dashboard role chips, filled_by_role aggregates) uses the flipped
# "Advanced Rigger" / "Intermediate Rigger" / "Basic Rigger". Normalise SQL
# trade names into the canonical vocabulary so macro-derived filled counts
# merge cleanly with existing required_by_role keys.
MACRO_ROLE_RENAME = {
    "Rigger - Advanced":     "Advanced Rigger",
    "Rigger - Intermediate": "Intermediate Rigger",
    "Rigger - Basic":        "Basic Rigger",
}

# Compliance sheet (xll01 PersonnelCompetency) -> per-site dashboard short
# keys. Each per-site dashboard renders ticket columns using these short
# names; translating them here lets the ticket data drop straight into the
# dashboards' existing render code without mapping every consumer.
TICKET_MAP = {
    "Confined Spaces Entry":                              "cse",
    "Working at Heights":                                 "wah",
    "EWP":                                                "ewp",
    "CA-EBS - Compressed Air Emergency Breathing System": "ba",
    "LF - Forklift Truck":                                "fork",
    # HR Class = Heavy Rigid driver's licence. HRWL = High Risk Work Licence
    # (the legal register). Two different tickets — the dashboards label them
    # separately so conflating them under one key would hide which is held.
    "HR Class":                                           "hr",
    "HRWL":                                               "hrwl",
    "DG - Dogging":                                       "dog",
    "Gas Test Atmospheres":                               "gta",
    "First Aid":                                          "fa",
}
# Rigging is special: per-site dashboards expect a single `rig` field whose
# value is the level string ("Advanced" / "Intermediate" / "Basic"), not a
# boolean. Accept both the "RA/RI/RB -" shorthand and the "Rigger -" long form
# and collapse to the highest level held.
RIG_COMPS = {
    "RA - Advanced Rigging":     "Advanced",
    "Rigger - Advanced":         "Advanced",
    "RI - Intermediate Rigging": "Intermediate",
    "Rigger - Intermediate":     "Intermediate",
    "RB - Basic Rigging":        "Basic",
    "Rigger - Basic":            "Basic",
}
RIG_PRIORITY = {"Advanced": 3, "Intermediate": 2, "Basic": 1}
EXPIRING_SOON_DAYS = 30

# Module-level cache so the workbook only gets opened once per run, even when
# both parse_rapidcrews.build_shutdown() (to look up JobPlanningView required
# counts for RosterCut files) and shutdowns_from_macro_data() (to synthesise
# macro-only shutdowns) touch the same sheets.
_CACHE: dict | None = None

# (Client string, Site string from PersonnelRosterView) ->
#     (company_key, client_display_name, dashboard_site, project_label_base)
#
# Covalent and Tronox share the 'SOUTH WEST' client banner in the source
# system and are disambiguated by Site. The dashboard_site column is what
# gets rendered on the tiles — it deliberately differs from the source Site
# so "Covalent Lithium" reads as "Mt Holland" (the plant's location) to
# match the existing dashboard copy.
# Tuple: (company_key, client_display, dashboard_site, label_base, id_prefix)
# id_prefix overrides company_key when generating the shutdown id (company_key-YYYY-MM).
# Use it when a RosterCut-derived id uses a different prefix than company_key so
# the macro-only dedup in parse_rapidcrews.py correctly detects the collision.
CLIENT_SITE_MAP: dict[tuple[str, str], tuple[str, str, str, str, str]] = {
    ("SOUTH WEST", "Covalent Lithium"): ("covalent", "Covalent", "Mt Holland", "Mt Holland",      "covalent"),
    ("SOUTH WEST", "Tronox"):            ("tronox",   "Tronox",   "Kwinana",    "Major Shutdown",  "tronox"),
    ("CSBP",       "CSBP Kwinana"):      ("csbp",     "CSBP",     "Kwinana",    "CSBP Kwinana",    "csbp"),
    # Kleenheat rolls up under CSBP (WesCEF umbrella) but the RosterCut file
    # uses id prefix "kleenheat-", so we must match that to get correct dedup.
    ("SOUTH WEST", "Kleenheat"):         ("csbp",     "CSBP",     "Kwinana",    "KPF LNG Kleenheat", "kleenheat"),
}

# Schedule Types that count as "on the job" for start/end + crew_split
# purposes. Anything else ("Personal Event", "Annual Leave", "Working
# Elsewhere", ...) is a scheduling artefact, not actual attendance.
ONSITE_SCHED_TYPES = {"Day Shift", "Night Shift", "RNR"}
CREW_LABEL = {"Day Shift": "Day", "Night Shift": "Night", "RNR": "RNR"}

_MONTH_NAME = ("", "January", "February", "March", "April", "May", "June",
               "July", "August", "September", "October", "November", "December")


# --------------------------------------------------------------------------- sheet loaders

def _open() -> openpyxl.Workbook:
    return openpyxl.load_workbook(MACRO_FILE, data_only=True, read_only=True)


def _load_cache() -> dict:
    """Open the macro workbook once per run and cache every sheet we touch.

    Returned structure:
        {
          "active_jobnos":   set[int] | None,     # from ACTIVE_SHUTDOWNS; None if sheet missing
          "trades":          {tradeId -> displayName},
          "personnel":       {personnelId -> {name, role, mobile, hire_company}},
          "planning_all":    {jobNo -> {tradeName -> {"required": n, "filled": m}}},
          "resumes":         list[dict] | None,   # from RESUMES; None if sheet missing
        }

    `planning_all` covers EVERY JobNo in JobPlanningView (not just the ones in
    ACTIVE_SHUTDOWNS) so parse_rapidcrews.build_shutdown() can look up required
    counts for any RosterCut JobNo that happens to be in the live SQL view.
    """
    global _CACHE
    if _CACHE is not None:
        return _CACHE
    if not MACRO_FILE.exists():
        _CACHE = {"active_jobnos": None, "active_labels": {}, "trades": {},
                  "personnel": {}, "personnel_name_index": {},
                  "planning_all": {}, "compliance": {}, "resumes": None}
        return _CACHE
    wb = _open()
    try:
        active_jobnos          = _read_active_shutdowns(wb)
        active_labels          = _read_active_shutdown_labels(wb)
        trades                 = _load_trade_names(wb)
        personnel              = _load_personnel(wb)
        personnel_name_index   = _load_personnel_name_index(wb)
        planning_all           = _load_planning_all(wb, trades)
        compliance             = _load_compliance(wb)
    finally:
        wb.close()
    resumes = _load_resumes_file(personnel) if RESUMES_FILE.exists() else None
    _CACHE = {
        "active_jobnos":        active_jobnos,
        "active_labels":        active_labels,
        "trades":               trades,
        "personnel":            personnel,
        "personnel_name_index": personnel_name_index,
        "planning_all":         planning_all,
        "compliance":           compliance,
        "resumes":              resumes,
    }
    return _CACHE


def _read_active_shutdowns(wb: openpyxl.Workbook) -> set[int] | None:
    if CONTROL_SHEET not in wb.sheetnames:
        return None
    ws = wb[CONTROL_SHEET]
    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        return set()
    try:
        jobno_col = [h for h in header if h].index("JobNo")
    except ValueError:
        print(f"  warn: {CONTROL_SHEET} sheet is missing 'JobNo' header")
        return set()
    jobnos: set[int] = set()
    for row in rows:
        if not row or row[jobno_col] is None:
            continue
        try:
            jobnos.add(int(row[jobno_col]))
        except (TypeError, ValueError):
            print(f"  warn: {CONTROL_SHEET} ignoring non-numeric JobNo {row[jobno_col]!r}")
    return jobnos


def _read_active_shutdown_labels(wb: openpyxl.Workbook) -> dict[int, str]:
    """JobNo -> descriptive label from ACTIVE_SHUTDOWNS sheet (col 2).

    The control sheet pairs each JobNo with a free-text description like
    "CSBP NaaN1" / "CSBP Naan2" / "Tronox May 2026". When two ACTIVE jobs
    map to the same id_prefix-YYYY-MM (e.g. NAAN1 and NAAN2 both starting
    in May 2026), this label disambiguates them in the project name so the
    dashboard's shutdown card and gantt swimlane don't both read identical.
    """
    if CONTROL_SHEET not in wb.sheetnames:
        return {}
    ws = wb[CONTROL_SHEET]
    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        return {}
    cols = [h for h in header if h]
    try:
        jobno_col = cols.index("JobNo")
    except ValueError:
        return {}
    out: dict[int, str] = {}
    for row in rows:
        if not row or row[jobno_col] is None:
            continue
        try:
            jno = int(row[jobno_col])
        except (TypeError, ValueError):
            continue
        # Take the next non-empty cell as the description.
        for cell in row[jobno_col + 1:]:
            text = str(cell or "").strip()
            if text:
                out[jno] = text
                break
    return out


def _load_planning_all(wb: openpyxl.Workbook,
                       trade_names: dict[str, str]
                       ) -> dict[int, dict[str, dict[str, int]]]:
    """JobNo -> {tradeName -> {"required": n, "filled": m}} for every JobNo."""
    ws = wb[JOB_PLANNING_SHEET]
    headers = [c.value for c in next(ws.iter_rows(max_row=1))]
    idx = {h: i for i, h in enumerate(headers) if h}
    out: dict[int, dict[str, dict[str, int]]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        job = row[idx["JobNo"]]
        if job is None:
            continue
        trade = trade_names.get(row[idx["CompetencyId"]], "Unknown")
        bucket = out.setdefault(int(job), {})
        cell   = bucket.setdefault(trade, {"required": 0, "filled": 0})
        cell["required"] += int(row[idx["Required"]] or 0)
        cell["filled"]   += int(row[idx["Filled"]]   or 0)
    return out


def planning_required_for_jobno(job_no: int) -> dict[str, int] | None:
    """Return JobPlanningView-derived `{tradeName: required}` for a JobNo.

    None means either the macro file doesn't exist OR the JobNo isn't in the
    live SQL view (dropped because the date range rolled over). Empty dict
    means the JobNo exists but all Required columns are zero.
    """
    cache = _load_cache()
    bucket = cache["planning_all"].get(int(job_no))
    if bucket is None:
        return None
    # Drop trades with 0 required AND 0 filled — stale placeholder rows.
    return {
        trade: cell["required"]
        for trade, cell in bucket.items()
        if cell["required"] or cell["filled"]
    }


def planning_filled_for_jobno(job_no: int) -> dict[str, int] | None:
    """Return JobPlanningView-derived `{tradeName: filled}` for a JobNo.

    This is the authoritative "slot filled" count — the same figure the
    Rapid Crews website shows on its planning page. Paired with
    `planning_required_for_jobno` so a live shutdown's headline numbers
    match the website without anyone needing to re-export a RosterCut.
    """
    cache = _load_cache()
    bucket = cache["planning_all"].get(int(job_no))
    if bucket is None:
        return None
    return {
        trade: cell["filled"]
        for trade, cell in bucket.items()
        if cell["required"] or cell["filled"]
    }


def active_shutdowns_jobnos() -> set[int] | None:
    """Return the set of JobNos from the ACTIVE_SHUTDOWNS control sheet.

    Returns None if the sheet doesn't exist at all (legacy mode — every
    RosterCut file in data/raw/ passes through). Returns a set (possibly
    empty) when the sheet is present, signalling the allow-list is active.
    """
    return _load_cache()["active_jobnos"]


def _load_trade_names(wb: openpyxl.Workbook) -> dict[str, str]:
    """TradeId GUID -> Trade display name (from DisciplineTrade).

    Trade names pass through MACRO_ROLE_RENAME so SQL's "Rigger - Advanced"
    becomes the "Advanced Rigger" key the rest of the pipeline uses.
    """
    ws = wb[TRADE_SHEET]
    headers = [c.value for c in next(ws.iter_rows(max_row=1))]
    idx = {h: i for i, h in enumerate(headers)}
    out: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        tid = row[idx["TradeId"]]
        name = (row[idx["Trade"]] or "").strip() or "Unknown"
        if tid:
            out[tid] = MACRO_ROLE_RENAME.get(name, name)
    return out


def _load_personnel(wb: openpyxl.Workbook) -> dict[str, dict]:
    """Personnel Id GUID -> {name, role, mobile, hire_company}."""
    ws = wb[PERSONNEL_SHEET]
    headers = [c.value for c in next(ws.iter_rows(max_row=1))]
    idx = {h: i for i, h in enumerate(headers)}
    out: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        pid = row[idx["Personnel Id"]]
        if not pid:
            continue
        first = (row[idx["Given Names"]] or "").strip()
        last  = (row[idx["Surname"]]      or "").strip()
        out[pid] = {
            "name":         f"{first} {last}".strip() or "Unknown",
            "role":         (row[idx["Primary Role"]] or "Unknown").strip(),
            "mobile":       rc._standardise_mobile(row[idx["Mobile"]]),
            "hire_company": (row[idx["Hire Company"]] or "").strip(),
        }
    return out


# --------------------------------------------------------------------------- compliance (tickets)

def _norm_name_part(s) -> str:
    """Aggressive first/last-name normalisation: lowercase, letters only.
    Drops spaces, hyphens, apostrophes so "O'Brien"/"OBrien" collide,
    "Van Der Zanden"/"VANDERZANDEN" too."""
    import re as _re
    return _re.sub(r"[^a-z]+", "", (s or "").lower())


def _load_personnel_name_index(wb: openpyxl.Workbook) -> dict[tuple[str, str], str]:
    """Build {(norm_first, norm_last): personnel_id} from xll01 Personnel.

    Personnel can have multiple rows (re-hires, profile duplicates). Later
    rows overwrite earlier under the same key; the compliance loader
    aggregates tickets across all of a person's IDs when matched.
    """
    if PERSONNEL_SHEET not in wb.sheetnames:
        return {}
    ws = wb[PERSONNEL_SHEET]
    headers = list(next(ws.iter_rows(max_row=1, values_only=True)))
    idx = {h: n for n, h in enumerate(headers) if h}
    out: dict[tuple[str, str], str] = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        pid = r[idx["Personnel Id"]]
        if not pid:
            continue
        f = _norm_name_part(r[idx["Given Names"]])
        l = _norm_name_part(r[idx["Surname"]])
        if f and l:
            out[(f, l)] = pid
    return out


def _load_compliance(wb: openpyxl.Workbook) -> dict[str, dict[str, dict]]:
    """Read xll01 PersonnelCompetency -> {pid: {ticket_key: record}}.

    Record shape: {"expiry": iso|None, "doc": url|None,
                   "status": "current"|"expiring_soon",
                   "level": str (rig only)}

    Filtering: skip Archived rows (superseded) and expired rows (Expiry ≤
    today). Rows with no Expiry are kept as permanent certs.

    Collisions:
      - Non-rig: keep the record with the LATEST expiry (permanent > dated).
      - Rig:     keep the HIGHEST level held (Advanced > Intermediate > Basic).
    """
    if COMPLIANCE_SHEET not in wb.sheetnames:
        return {}
    ws = wb[COMPLIANCE_SHEET]
    headers = list(next(ws.iter_rows(max_row=1, values_only=True)))
    idx = {h: n for n, h in enumerate(headers) if h}
    today = dt.date.today()
    soon  = today + dt.timedelta(days=EXPIRING_SOON_DAYS)
    out: dict[str, dict[str, dict]] = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[idx["Archived"]]:
            continue
        comp = r[idx["Competency"]]
        pid  = r[idx["Personnel Id"]]
        if not comp or not pid:
            continue
        is_rig = comp in RIG_COMPS
        key    = "rig" if is_rig else TICKET_MAP.get(comp)
        if key is None:
            continue
        exp = r[idx["Expiry"]]
        exp_date: dt.date | None = None
        if exp:
            exp_date = exp.date() if isinstance(exp, dt.datetime) else exp
            if not isinstance(exp_date, dt.date):
                exp_date = None
            elif exp_date <= today:
                continue
        record = {
            "expiry": exp_date.isoformat() if exp_date else None,
            "doc":    r[idx["Document Location"]] or None,
            "status": "expiring_soon" if exp_date and exp_date <= soon else "current",
        }
        if is_rig:
            record["level"] = RIG_COMPS[comp]
            existing = out.setdefault(pid, {}).get(key)
            if existing is None or RIG_PRIORITY[record["level"]] > RIG_PRIORITY[existing["level"]]:
                out[pid][key] = record
        else:
            existing = out.setdefault(pid, {}).get(key)
            if existing is None:
                out[pid][key] = record
            else:
                e_exp = existing["expiry"]
                if record["expiry"] is None and e_exp is not None:
                    out[pid][key] = record
                elif record["expiry"] is not None and e_exp is not None and record["expiry"] > e_exp:
                    out[pid][key] = record
    return out


def match_personnel_id(first: str, last: str) -> str | None:
    """Resolve a (first, last) name pair to a Personnel Id via three passes:
    exact normalised match, prefix-on-first-name-with-same-surname (handles
    "Lucrecia" vs "Lucrecia Celeste"), then surname-unique fallback.
    Returns None when no personnel master data is loaded."""
    cache = _load_cache()
    index = cache.get("personnel_name_index") or {}
    f = _norm_name_part(first)
    l = _norm_name_part(last)
    if not f or not l:
        return None
    if (f, l) in index:
        return index[(f, l)]
    same_surname = [(if_, pid) for (if_, il_), pid in index.items() if il_ == l]
    for if_, pid in same_surname:
        if if_.startswith(f) or f.startswith(if_):
            return pid
    if len(same_surname) == 1:
        return same_surname[0][1]
    return None


def tickets_for_person(first: str, last: str) -> dict:
    """Return the tickets dict for a worker, or {} when no match / no data.
    Safe to call even when the macro workbook or compliance sheet is
    missing — returns empty."""
    pid = match_personnel_id(first, last)
    if not pid:
        return {}
    cache = _load_cache()
    return (cache.get("compliance") or {}).get(pid, {})


def _load_resumes_file(personnel: dict[str, dict]) -> list[dict]:
    """Read the standalone Resumes.xlsx file into a flat list of dicts.

    End-user schema (all columns optional except Name OR Personnel Id):
        Name | Personnel Id | Role | Mobile | Resume URL | Updated | Notes

    Resume URL is typically a SharePoint share link. If Personnel Id is set
    we merge in mobile/role from xll01 Personnel where the sheet leaves them
    blank, so the handover form stays minimal ("just paste the link").

    Kept in its own xlsx rather than as a sheet inside the macro workbook
    because: (a) ops owns it and shouldn't have to touch the Rapid Crews SQL
    export, (b) SharePoint sync pulls are lighter when resume updates don't
    require re-uploading a 15MB workbook.
    """
    wb = openpyxl.load_workbook(RESUMES_FILE, data_only=True, read_only=True)
    try:
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            return []
        hdr = [(h or "").strip() for h in header]
        # Accept both "Resume URL" and "Resume Link" — humans type either.
        def col(*names):
            for n in names:
                if n in hdr:
                    return hdr.index(n)
            return None
        c_name   = col("Name", "Worker")
        c_pid    = col("Personnel Id", "PersonnelId", "Employee ID")
        c_role   = col("Role", "Primary Role", "Trade")
        c_mobile = col("Mobile", "Phone", "Contact")
        c_url    = col("Resume URL", "Resume Link", "URL", "Link")
        c_upd    = col("Updated", "Updated Date", "Last Updated", "Date")
        c_notes  = col("Notes", "Comment")
        out: list[dict] = []
        for row in rows:
            if not row or not any(row):
                continue
            pid  = (row[c_pid]  if c_pid  is not None else None) or ""
            name = (row[c_name] if c_name is not None else None) or ""
            p    = personnel.get(pid) if pid else None
            def pick(col_idx, fallback):
                v = row[col_idx] if col_idx is not None else None
                s = str(v).strip() if v is not None else ""
                return s or (fallback or "")
            # Name: sheet wins, else look up from personnel master by GUID.
            disp_name = str(name).strip() or (p["name"] if p else "")
            if not disp_name:
                continue
            out.append({
                "name":         disp_name,
                "personnel_id": str(pid).strip() or None,
                "role":         pick(c_role, p["role"] if p else ""),
                "mobile":       rc._standardise_mobile(
                                    row[c_mobile] if c_mobile is not None else None
                                ) or (p["mobile"] if p else ""),
                "resume_url":   pick(c_url, ""),
                "updated":      pick(c_upd, ""),
                "notes":        pick(c_notes, ""),
            })
        return out
    finally:
        wb.close()


def resumes_from_macro_data() -> list[dict]:
    """Public entry point — returns parsed rows from Resumes.xlsx (or empty
    list when the file is absent). Dashboard cross-references `resume_url`
    by normalised name to decorate worker rows with a link."""
    return _load_cache()["resumes"] or []


def _load_roster(wb: openpyxl.Workbook,
                 jobnos: set[int]
                 ) -> dict[int, dict]:
    """JobNo -> {client, site, workers: {PersonnelId -> {dates, sched_types,
    is_on_location}}}. Daily schedule rows are folded into one entry per
    (JobNo, Personnel Id)."""
    ws = wb[ROSTER_VIEW_SHEET]
    headers = [c.value for c in next(ws.iter_rows(max_row=1))]
    idx = {h: i for i, h in enumerate(headers)}
    out: dict[int, dict] = {
        j: {"client": None, "site": None, "workers": defaultdict(lambda: {
            "dates": [], "sched_types": Counter(), "is_on_location": False,
        })}
        for j in jobnos
    }
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        job = row[idx["Job No"]]
        if job not in out:
            continue
        bucket = out[job]
        if bucket["client"] is None:
            bucket["client"] = row[idx["Client"]]
            bucket["site"]   = row[idx["Site"]]
        pid = row[idx["Personnel Id"]]
        if not pid:
            continue
        w = bucket["workers"][pid]
        date = row[idx["Schedule Date"]]
        if isinstance(date, dt.datetime):
            w["dates"].append(date.date())
        elif isinstance(date, dt.date):
            w["dates"].append(date)
        sched = row[idx["Schedule Type"]]
        if sched:
            w["sched_types"][sched] += 1
        if row[idx["IsOnLocation"]]:
            w["is_on_location"] = True
    return out


def _load_daily_personnel_schedule(wb: openpyxl.Workbook,
                                   jobnos: set[int]
                                   ) -> dict[int, dict]:
    """JobNo -> {PersonnelId -> {role, statuses, status_latest}}.

    DailyPersonnelSchedule has one row per (person, job, day) with a Status
    field that tracks the worker's placement state on that job (Contacted →
    Short List → Confirmed → Mobilising → Onsite → Demobilised, with
    Declined / Late Withdrawal / Rejected as off-ramps). This loader keeps
    every (job, pid) the macro file mentions for any active job, so callers
    can decide which Status set to treat as "filled". `role` is taken from
    the Trade column and renamed to the canonical vocabulary.

    A worker can have rows in multiple statuses on the same day (e.g.
    "Onsite" + "Rejected" appearing together when a roster slot is
    reassigned). For the displayed `latest_status` we pick the most-recent
    row whose status is in ONSITE_PERSONNEL_STATUSES so a single trailing
    "Rejected" doesn't paper over a real on-site stint.
    """
    if DAILY_SCHEDULE_SHEET not in wb.sheetnames:
        # _build_one falls back to PRV-only logic when DPS data is absent,
        # which still works but loses the status filter (Confirmed /
        # Mobilising / Onsite / Demobilised). Surface the missing sheet so
        # users notice they're running against an old workbook export.
        print(f"  warn: macro workbook is missing the {DAILY_SCHEDULE_SHEET!r} "
              f"sheet — falling back to PersonnelRosterView without a status "
              f"filter. Re-export the workbook to restore named-roster filtering.")
        return {j: {} for j in jobnos}
    ws = wb[DAILY_SCHEDULE_SHEET]
    headers = [c.value for c in next(ws.iter_rows(max_row=1))]
    idx = {h: i for i, h in enumerate(headers) if h}
    out: dict[int, dict] = {j: {} for j in jobnos}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        job = row[idx["JobId"]]
        if job not in out:
            continue
        pid = row[idx["PersonnelId"]]
        if not pid:
            continue
        status = (row[idx["Status"]] or "").strip()
        trade  = (row[idx["Trade"]]  or "").strip()
        trade  = MACRO_ROLE_RENAME.get(trade, trade)
        report = row[idx["ReportDate"]]
        if isinstance(report, dt.datetime):
            report = report.date()
        rec = out[job].get(pid)
        if rec is None:
            rec = {
                "trades":         Counter(),
                "statuses":       set(),
                "report_dates":   [],
                "onsite_days":    0,
                "latest_status":  ("", None),  # (status, date) within ONSITE set only
            }
            out[job][pid] = rec
        if status:
            rec["statuses"].add(status)
            if status in ONSITE_PERSONNEL_STATUSES:
                cur_status, cur_date = rec["latest_status"]
                if report and (cur_date is None or report > cur_date):
                    rec["latest_status"] = (status, report)
        if trade and status in ONSITE_PERSONNEL_STATUSES:
            rec["trades"][trade] += 1
        if report and status in ONSITE_PERSONNEL_STATUSES:
            rec["report_dates"].append(report)
        if row[idx.get("OnSite")] in (1, True):
            rec["onsite_days"] += 1
    return out


# --------------------------------------------------------------------------- builder

def _project_label(base: str, start: dt.date) -> str:
    return f"{base} {_MONTH_NAME[start.month]} {start.year}"


def _build_one(job_no: int,
               planning: dict,
               roster_raw: dict,
               personnel: dict[str, dict]) -> tuple[str, str, dict] | None:
    client = roster_raw["client"]
    site   = roster_raw["site"]
    if (client, site) not in CLIENT_SITE_MAP:
        print(f"  warn: JobNo {job_no} has unmapped client/site "
              f"({client!r}, {site!r}) — skipping. Add to CLIENT_SITE_MAP.")
        return None
    company_key, client_name, dashboard_site, label_base, id_prefix = CLIENT_SITE_MAP[(client, site)]

    workers   = roster_raw["workers"]
    dps       = roster_raw.get("dps", {}) or {}
    # The named roster is the union of (a) PersonnelRosterView entries with at
    # least one on-site schedule day and (b) DailyPersonnelSchedule entries
    # whose Status is Confirmed/Mobilising/Onsite/Demobilised. In practice
    # both lists are the same set of PIDs for active shutdowns — but we
    # accept either side so a worker confirmed in DPS without a daily roster
    # row yet still appears, and an on-site PRV worker without a DPS row
    # still appears (legacy completed jobs where DPS has aged out).
    pids = set(workers) | set(dps)
    worker_candidates: list[dict] = []
    filled_by_role: Counter = Counter()
    skipped_dps_unfilled = 0
    for pid in pids:
        w   = workers.get(pid) or {"dates": [], "sched_types": Counter(), "is_on_location": False}
        dps_rec = dps.get(pid)

        # Status filter: when DPS knows about this person on this job, only
        # keep them if they're in an "actually filling a slot" status. PRV-
        # only PIDs (legacy / older snapshots without DPS coverage) keep the
        # previous "any on-site Schedule Type" rule so we don't drop history.
        if dps_rec is not None:
            if not (dps_rec["statuses"] & ONSITE_PERSONNEL_STATUSES):
                skipped_dps_unfilled += 1
                continue
        else:
            onsite_dates = [d for d, st in zip(w["dates"],
                                               _explode_sched_types(w["sched_types"], len(w["dates"])))
                            if st in ONSITE_SCHED_TYPES]
            if not onsite_dates:
                continue

        person = personnel.get(pid)
        if not person:
            continue
        mobile = person["mobile"]

        # Per-job role: prefer DPS Trade (the trade this person is filling
        # *on this job* — what JobPlanningView counts against). Fall back to
        # the worker's Primary Role from xll01 Personnel when DPS doesn't
        # cover this PID.
        if dps_rec and dps_rec["trades"]:
            role = dps_rec["trades"].most_common(1)[0][0]
        else:
            role = person["role"]

        # Date span — prefer PRV scheduled days when present, else DPS
        # ReportDates filtered to ONSITE statuses.
        if w["dates"]:
            start = min(w["dates"])
            end   = max(w["dates"])
        elif dps_rec and dps_rec["report_dates"]:
            start = min(dps_rec["report_dates"])
            end   = max(dps_rec["report_dates"])
        else:
            # Only happens for PIDs in DPS with onsite status but no report
            # dates yet — treat as open-ended at today.
            start = end = dt.date.today()

        # Dominant on-site Schedule Type drives the crew label. DPS doesn't
        # carry shift labels, so PRV-less PIDs default to "Day".
        onsite_only = Counter({k: v for k, v in w["sched_types"].items()
                               if k in ONSITE_SCHED_TYPES})
        if onsite_only:
            dom_sched  = onsite_only.most_common(1)[0][0]
            crew_label = CREW_LABEL.get(dom_sched, dom_sched)
        else:
            crew_label = "Day"

        entry = {
            "name":         person["name"],
            "role":         role,
            "shift":        crew_label,
            "start":        start.isoformat(),
            "end":          end.isoformat(),
            "personnel_id": pid,
            "tickets":      (_load_cache().get("compliance") or {}).get(pid, {}),
        }
        if mobile:
            entry["mobile"] = mobile
        if dps_rec and dps_rec["latest_status"][0]:
            entry["status"] = dps_rec["latest_status"][0]

        worker_candidates.append({
            "entry":          entry,
            "role":           role,
            "shift":          crew_label,
            "hire_company":   person["hire_company"],
            "is_on_location": bool(w["is_on_location"]) or bool(dps_rec and dps_rec["onsite_days"]),
            "onsite_days":    max(len([d for d, st in zip(w["dates"],
                                                          _explode_sched_types(w["sched_types"], len(w["dates"])))
                                       if st in ONSITE_SCHED_TYPES]),
                                  (dps_rec["onsite_days"] if dps_rec else 0)),
        })
        filled_by_role[role] += 1

    if skipped_dps_unfilled:
        print(f"  macro: JobNo {job_no} skipped {skipped_dps_unfilled} workers "
              f"with non-filling DPS status (Contacted/Declined/Rejected/etc.)")
    if not worker_candidates:
        print(f"  warn: JobNo {job_no} has no on-site workers — skipping")
        return None

    # DailyPersonnelSchedule already gave us the authoritative filled list
    # above. No JobPlanningView cap — we used to chop roster rows down to
    # JP's per-role Filled count, but that hid people RapidCrews considered
    # mobilising/on-site. The named roster IS the Filled list now, and
    # filled_by_role is derived from it directly.
    planning_req = planning["required_by_role"]
    planning_fil = planning["filled_by_role"]
    roster_selected = list(worker_candidates)

    if not roster_selected:
        print(f"  warn: JobNo {job_no} has no selected workers — skipping")
        return None

    # Shutdown window = span of selected workers' earliest/latest dates.
    all_starts = [dt.date.fromisoformat(c["entry"]["start"]) for c in roster_selected]
    all_ends   = [dt.date.fromisoformat(c["entry"]["end"])   for c in roster_selected]
    sd, ed = min(all_starts), max(all_ends)
    shutdown_id = f"{id_prefix}-{sd.isoformat()[:7]}"

    roster_entries = sorted((c["entry"] for c in roster_selected), key=lambda r: r["name"])
    crew_split: Counter = Counter(c["shift"] for c in roster_selected)
    mobilised_by_role: Counter = Counter(c["role"] for c in roster_selected if c["is_on_location"])
    labour_hire_split: Counter = Counter(c["hire_company"] for c in roster_selected)

    if planning_req or any(planning_fil.values()):
        # filled_by_role is derived from the named roster (one row per unique
        # PersonnelId, role = DPS Trade), so the per-role table sums to the
        # same total that the matrix and KPI cards show. Required keeps using
        # JobPlanningView (the only place the full demand by trade lives).
        # planning_fil is preserved on _source so the dashboard can flag the
        # gap when JobPlanningView and the named roster disagree.
        selected_by_role = Counter(c["role"] for c in roster_selected)
        all_keys = set(selected_by_role) | set(planning_req) | set(planning_fil)
        required     = {r: int(planning_req.get(r, 0))     for r in all_keys}
        filled_final = {r: int(selected_by_role.get(r, 0)) for r in all_keys}
        target_source_meta     = {"source": "rapid_crews_job_planning_view",
                                  "job_no": job_no,
                                  "total_required":          sum(planning_req.values()),
                                  "total_filled_planning":   sum(planning_fil.values()),
                                  "total_filled_named":      len(roster_selected)}
        required_target_source = "RAPID_CREWS_JOB_PLANNING"
    else:
        required, filled_final, target_source_meta = rc.merge_targets(shutdown_id, filled_by_role)
        required_target_source = (
            "TARGET_FILE" if (rc.TARGETS_DIR / f"{shutdown_id}.json").exists()
            else "PLACEHOLDER_FROM_ROSTER"
        )

    status = rc._infer_status(sd, ed, dt.date.today())

    # Prefer the descriptive label from the ACTIVE_SHUTDOWNS sheet (e.g.
    # "CSBP NaaN1" / "CSBP Naan2") when supplied — disambiguates two
    # shutdowns at the same site in the same month. Fall back to the
    # generic site-month template. Some labels in the sheet already
    # include the month/year ("Tronox May 2026"); only append when missing
    # so we don't end up with "Tronox May 2026 May 2026".
    label_from_sheet = (_load_cache().get("active_labels") or {}).get(int(job_no))
    if label_from_sheet:
        month = _MONTH_NAME[sd.month]
        year  = str(sd.year)
        if month in label_from_sheet and year in label_from_sheet:
            project_name = label_from_sheet
        else:
            project_name = f"{label_from_sheet} {month} {year}"
    else:
        project_name = _project_label(label_base, sd)

    shutdown = {
        "id":               shutdown_id,
        "name":             project_name,
        "site":             dashboard_site,
        "start_date":       sd.isoformat(),
        "end_date":         ed.isoformat(),
        "status":           status,
        "required_by_role": required,
        "filled_by_role":   filled_final,
        "crew_split":       dict(crew_split),
        "mobilised_by_role": dict(mobilised_by_role),
        "labour_hire_split": dict(labour_hire_split),
        "roster":           sorted(roster_entries, key=lambda r: r["name"]),
        "_source": {
            "macro_data_job_no":           job_no,
            "macro_data_client":           client,
            "macro_data_site":             site,
            "source_format":               "macro_data",
            "required_target_source":      required_target_source,
            "macro_data_roster_size":      len(roster_entries),
            "macro_data_filled_by_role":   dict(filled_by_role),
            # JobPlanningView's per-role Filled — preserved so the dashboard can
            # surface a "Planning view: N (M unnamed)" gap pill when it differs
            # from the named-roster total. Not used as a number on screen.
            "planning_filled_by_role":     dict(planning_fil),
            "target_source":               target_source_meta,
        },
    }
    return company_key, client_name, shutdown


def _explode_sched_types(counter: Counter, n_dates: int) -> list[str]:
    """Best-effort alignment between each schedule date and its Schedule
    Type for the worker. We only keep aggregate counts per type in the
    loader (simpler + lower memory), so reconstruct a per-date sequence
    by repeating each type by its count. Order won't match date order,
    but for the onsite filter we only care about set membership per
    worker, which this preserves."""
    seq: list[str] = []
    for st, c in counter.items():
        seq.extend([st] * c)
    # Pad/truncate to requested length (guards against divergence from
    # len(dates) when a row had no Schedule Type at all).
    if len(seq) < n_dates:
        seq.extend([None] * (n_dates - len(seq)))
    return seq[:n_dates]


# --------------------------------------------------------------------------- public entry point

def shutdowns_from_macro_data() -> list[tuple[str, str, dict]]:
    """Return (company_key, client_name, shutdown_dict) triples for every
    JobNo in the ACTIVE_SHUTDOWNS sheet. Empty list if sheet is absent or
    empty, if the macro file doesn't exist, or if no JobNo resolves to a
    mapped client/site + non-empty roster."""
    if not MACRO_FILE.exists():
        return []
    cache = _load_cache()
    jobnos = cache["active_jobnos"]
    if not jobnos:
        return []
    print(f"  macro: ACTIVE_SHUTDOWNS lists {len(jobnos)} JobNo(s): "
          f"{sorted(jobnos)}")

    # PersonnelRosterView gives us per-day Schedule Type (Day/Night/RNR) for
    # crew_split. DailyPersonnelSchedule is the source of truth for who's
    # actually filling a slot on each job (Status: Confirmed/Mobilising/
    # Onsite/Demobilised) and what trade they're filling it as. We pull both
    # and combine in _build_one so the named roster matches the matrix and
    # the shutdown-detail filled count.
    wb = _open()
    try:
        roster = _load_roster(wb, jobnos)
        dps    = _load_daily_personnel_schedule(wb, jobnos)
    finally:
        wb.close()
    for job, dps_workers in dps.items():
        if job in roster:
            roster[job]["dps"] = dps_workers

    # Re-shape the cached planning rows into the legacy {required_by_role,
    # filled_by_role} shape _build_one expects.
    planning: dict[int, dict] = {}
    for job in jobnos:
        bucket = cache["planning_all"].get(int(job), {})
        planning[job] = {
            "required_by_role": {t: c["required"] for t, c in bucket.items()
                                 if c["required"] or c["filled"]},
            "filled_by_role":   {t: c["filled"]   for t, c in bucket.items()
                                 if c["required"] or c["filled"]},
        }

    out: list[tuple[str, str, dict]] = []
    for job in sorted(jobnos):
        bucket = roster.get(job)
        has_prv = bool(bucket and bucket["workers"])
        has_dps = bool(bucket and bucket.get("dps"))
        if not (has_prv or has_dps):
            print(f"  warn: JobNo {job} has no roster rows in macro data — skipping")
            continue
        result = _build_one(job,
                            planning.get(job, {"required_by_role": {}, "filled_by_role": {}}),
                            bucket, cache["personnel"])
        if result is not None:
            out.append(result)
            _, client_name, shutdown = result
            print(f"  macro: JobNo {job:>4} -> {client_name:<10} "
                  f"{shutdown['id']:<22} roster={len(shutdown['roster']):>3}  "
                  f"{shutdown['start_date']} → {shutdown['end_date']}  "
                  f"[{shutdown['status']}]")
    return out


if __name__ == "__main__":
    # Ad-hoc: print what the macro parser would emit, no JSON write.
    for ck, cn, s in shutdowns_from_macro_data():
        print(ck, cn, s["id"], s["name"], s["start_date"], s["end_date"],
              f"roster={len(s['roster'])}")
