#!/usr/bin/env python3
"""Apply Hiring Company data from Rapidcrews Macro Data.xlsx to dashboard JSON.

The dashboard roster/matrix needs to show where each worker is coming from.
RapidCrews carries this in the macro workbook's personnel and roster views;
this pass adds `hire_company` onto each roster entry in data/*.json and
matching data/history/*.json snapshots.

This script is deliberately defensive: missing sheets, missing columns or an
unreadable workbook should not fail the whole dashboard refresh.
"""
from __future__ import annotations

import json
import pathlib
import re
import sys
from typing import Any

import openpyxl

REPO_ROOT = pathlib.Path(__file__).resolve().parent.parent
DATA_DIR = REPO_ROOT / "data"
HISTORY_DIR = DATA_DIR / "history"
MACRO_FILE = DATA_DIR / "raw" / "Rapidcrews Macro Data.xlsx"
PERSONNEL_SHEET = "xll01 Personnel"
ROSTER_VIEW_SHEET = "xpbi02 PersonnelRosterView"


def _clean(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\xa0", " ")).strip()


def _name_key(*parts: Any) -> str:
    text = " ".join(_clean(p) for p in parts if _clean(p))
    return re.sub(r"[^a-z]+", "", text.lower())


def _headers(ws) -> dict[str, int]:
    try:
        header = list(next(ws.iter_rows(max_row=1, values_only=True)))
    except StopIteration:
        return {}
    return {_clean(h): i for i, h in enumerate(header) if _clean(h)}


def _first_idx(idx: dict[str, int], *names: str) -> int | None:
    lower = {k.lower(): v for k, v in idx.items()}
    for name in names:
        if name in idx:
            return idx[name]
        if name.lower() in lower:
            return lower[name.lower()]
    return None


def _read_rows_into_maps(ws, by_pid: dict[str, str], by_name: dict[str, str]) -> None:
    idx = _headers(ws)
    if not idx:
        return

    pid_i = _first_idx(idx, "Personnel Id", "PersonnelId", "Personnel ID")
    hire_i = _first_idx(idx, "Hire Company", "Hiring Company", "Company", "Labour Hire", "Labour Hire Company")
    if hire_i is None:
        return

    first_i = _first_idx(idx, "Given Names", "Given Name", "First Name", "Name", "Personnel")
    last_i = _first_idx(idx, "Surname", "Last Name", "Last")

    for row_no, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            if not row or not any(row):
                continue
            hire = _clean(row[hire_i] if hire_i < len(row) else "")
            if not hire:
                continue

            if pid_i is not None and pid_i < len(row):
                pid = _clean(row[pid_i])
                if pid:
                    by_pid[pid] = hire

            first = _clean(row[first_i]) if first_i is not None and first_i < len(row) else ""
            last = _clean(row[last_i]) if last_i is not None and last_i < len(row) else ""

            # Some views carry one full-name field rather than first/surname.
            key = _name_key(first, last)
            if key:
                by_name.setdefault(key, hire)
        except Exception as exc:
            print(f"::warning::apply_hiring_company skipped row {row_no} in {ws.title}: {exc}", file=sys.stderr)


def _read_hire_company_maps() -> tuple[dict[str, str], dict[str, str]]:
    """Return (personnel_id -> hire_company, normalised_name -> hire_company)."""
    if not MACRO_FILE.exists():
        print("apply_hiring_company: no macro workbook found; skipped")
        return {}, {}

    by_pid: dict[str, str] = {}
    by_name: dict[str, str] = {}

    try:
        wb = openpyxl.load_workbook(MACRO_FILE, data_only=True, read_only=True)
    except Exception as exc:
        print(f"::warning::apply_hiring_company could not open macro workbook: {exc}", file=sys.stderr)
        return {}, {}

    try:
        for sheet_name in (PERSONNEL_SHEET, ROSTER_VIEW_SHEET):
            if sheet_name not in wb.sheetnames:
                print(f"apply_hiring_company: sheet not found: {sheet_name}")
                continue
            try:
                _read_rows_into_maps(wb[sheet_name], by_pid, by_name)
            except Exception as exc:
                print(f"::warning::apply_hiring_company skipped sheet {sheet_name}: {exc}", file=sys.stderr)
    finally:
        wb.close()

    print(f"apply_hiring_company: loaded {len(by_pid)} personnel-id hire-company mappings and {len(by_name)} name mappings")
    return by_pid, by_name


def _apply_to_shutdown(shutdown: dict, by_pid: dict[str, str], by_name: dict[str, str]) -> bool:
    changed = False
    for worker in shutdown.get("roster", []) or []:
        if not isinstance(worker, dict):
            continue
        current = _clean(worker.get("hire_company") or worker.get("hiring_company"))
        pid = _clean(worker.get("personnel_id"))
        hire = by_pid.get(pid, "") if pid else ""
        if not hire:
            hire = by_name.get(_name_key(worker.get("name")), "")
        if hire and current != hire:
            worker["hire_company"] = hire
            changed = True
    return changed


def _patch_company_file(path: pathlib.Path, by_pid: dict[str, str], by_name: dict[str, str]) -> bool:
    try:
        payload = json.loads(path.read_text())
    except Exception as exc:
        print(f"::warning::apply_hiring_company could not read {path.name}: {exc}", file=sys.stderr)
        return False
    changed = False
    for shutdown in payload.get("shutdowns", []) or []:
        changed = _apply_to_shutdown(shutdown, by_pid, by_name) or changed
    if changed:
        path.write_text(json.dumps(payload, indent=2))
    return changed


def _patch_history_file(path: pathlib.Path, by_pid: dict[str, str], by_name: dict[str, str]) -> bool:
    try:
        payload = json.loads(path.read_text())
    except Exception as exc:
        print(f"::warning::apply_hiring_company could not read {path.name}: {exc}", file=sys.stderr)
        return False
    shutdown = payload.get("shutdown")
    if not isinstance(shutdown, dict):
        return False
    changed = _apply_to_shutdown(shutdown, by_pid, by_name)
    if changed:
        path.write_text(json.dumps(payload, indent=2))
    return changed


def main() -> int:
    try:
        by_pid, by_name = _read_hire_company_maps()
        if not by_pid and not by_name:
            print("apply_hiring_company: no hiring-company mappings found; skipped")
            return 0

        changed_files: list[str] = []
        for name in ("covalent", "tronox", "csbp"):
            path = DATA_DIR / f"{name}.json"
            if path.exists() and _patch_company_file(path, by_pid, by_name):
                changed_files.append(str(path.relative_to(REPO_ROOT)))

        if HISTORY_DIR.exists():
            for path in sorted(HISTORY_DIR.glob("*.json")):
                if _patch_history_file(path, by_pid, by_name):
                    changed_files.append(str(path.relative_to(REPO_ROOT)))

        if changed_files:
            print("apply_hiring_company: updated")
            for f in changed_files:
                print(f"  - {f}")
        else:
            print("apply_hiring_company: no roster hire-company changes")
    except Exception as exc:
        print(f"::warning::apply_hiring_company failed defensively and was skipped: {exc}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
