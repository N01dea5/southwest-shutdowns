#!/usr/bin/env python3
"""Ensure every JobNo listed in Rapidcrews Macro Data.xlsx / ACTIVE_SHUTDOWNS
appears in the public dashboard JSON.

This is a safety pass for early-stage shutdowns. The main parser builds rich
shutdowns when it has usable roster rows. A newly added JobNo can legitimately
have planning demand before workers have on-site schedule rows; previously that
case was skipped, so the public dashboard did not show the new shutdown.

This script runs after scripts/parse_rapidcrews.py. It reads:
  - ACTIVE_SHUTDOWNS
  - xpbi02 JobPlanningView
  - xpbi02 DisciplineTrade
  - xpbi02 PersonnelRosterView
  - xll01 Personnel

For any active JobNo not already present in data/*.json, it creates a minimal
shutdown using planning counts and whatever roster metadata is available.
"""
from __future__ import annotations

import datetime as dt
import json
import pathlib
import re
from collections import Counter, defaultdict
from typing import Any

import openpyxl


REPO_ROOT = pathlib.Path(__file__).resolve().parent.parent
MACRO_FILE = REPO_ROOT / "data" / "raw" / "Rapidcrews Macro Data.xlsx"
DATA_DIR = REPO_ROOT / "data"
HISTORY_DIR = DATA_DIR / "history"

CONTROL_SHEET = "ACTIVE_SHUTDOWNS"
JOB_PLANNING_SHEET = "xpbi02 JobPlanningView"
ROSTER_VIEW_SHEET = "xpbi02 PersonnelRosterView"
TRADE_SHEET = "xpbi02 DisciplineTrade"
PERSONNEL_SHEET = "xll01 Personnel"

CLIENT_SITE_MAP: dict[tuple[str, str], tuple[str, str, str, str]] = {
    ("SOUTH WEST", "Covalent Lithium"): ("covalent", "Covalent", "Mt Holland", "Mt Holland"),
    ("SOUTH WEST", "Tronox"): ("tronox", "Tronox", "Kwinana", "Major Shutdown"),
    ("CSBP", "CSBP Kwinana"): ("csbp", "CSBP", "Kwinana", "CSBP Kwinana"),
    ("CSBP", "CSBP"): ("csbp", "CSBP", "Kwinana", "CSBP"),
    ("WESCEF", "CSBP Kwinana"): ("csbp", "CSBP", "Kwinana", "CSBP Kwinana"),
    ("WESCEF", "CSBP"): ("csbp", "CSBP", "Kwinana", "CSBP"),
    ("SOUTH WEST", "Kleenheat"): ("csbp", "CSBP", "Kwinana", "KPF LNG Kleenheat"),
}

ROLE_RENAME = {
    "Rigger - Advanced": "Advanced Rigger",
    "Rigger - Intermediate": "Intermediate Rigger",
    "Rigger - Basic": "Basic Rigger",
}

CREW_LABEL = {
    "Day Shift": "Day",
    "Night Shift": "Night",
    "RNR": "RNR",
}

MONTH_NAME = ("", "January", "February", "March", "April", "May", "June",
              "July", "August", "September", "October", "November", "December")


def _clean_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\xa0", " ")).strip()


def _map_client_site(client: Any, site: Any, job_no: int) -> tuple[str, str, str, str] | None:
    c = _clean_text(client)
    s = _clean_text(site)
    mapped = CLIENT_SITE_MAP.get((c, s))
    if mapped:
        return mapped
    cu = c.upper()
    su = s.upper()
    if "CSBP" in cu or "CSBP" in su or "NAAN" in su:
        return ("csbp", "CSBP", "Kwinana", "CSBP")
    if "TRONOX" in cu or "TRONOX" in su:
        return ("tronox", "Tronox", "Kwinana", "Major Shutdown")
    if "COVALENT" in cu or "COVALENT" in su or "MT HOLLAND" in su:
        return ("covalent", "Covalent", "Mt Holland", "Mt Holland")
    print(f"  warn: active JobNo {job_no} has unmapped client/site ({c!r}, {s!r}) — cannot emit dashboard card")
    return None


def _headers(ws) -> tuple[list[Any], dict[str, int]]:
    header = list(next(ws.iter_rows(max_row=1, values_only=True)))
    idx = {str(h).strip(): i for i, h in enumerate(header) if h is not None and str(h).strip()}
    return header, idx


def _to_date(value) -> dt.date | None:
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    if isinstance(value, str):
        text = value.strip()
        m = re.match(r"(\d{4}-\d{2}-\d{2})", text)
        if m:
            return dt.date.fromisoformat(m.group(1))
        for fmt in ("%d/%m/%y", "%d/%m/%Y", "%d-%m-%y", "%d-%m-%Y"):
            try:
                return dt.datetime.strptime(text, fmt).date()
            except ValueError:
                pass
    return None


def _status(start: dt.date, end: dt.date) -> str:
    today = dt.date.today()
    if end < today:
        return "completed"
    if start <= today:
        return "in_progress"
    return "booked"


def _mobile(raw) -> str:
    digits = re.sub(r"\D", "", str(raw or ""))
    if not digits:
        return ""
    if len(digits) == 11 and digits.startswith("61") and digits[2] == "4":
        digits = "0" + digits[2:]
    elif len(digits) == 9 and digits.startswith("4"):
        digits = "0" + digits
    if len(digits) == 10 and digits.startswith("04"):
        return f"{digits[0:4]} {digits[4:7]} {digits[7:10]}"
    return digits


def _read_active_jobnos(wb) -> set[int]:
    if CONTROL_SHEET not in wb.sheetnames:
        return set()
    ws = wb[CONTROL_SHEET]
    _, idx = _headers(ws)
    if "JobNo" not in idx:
        print(f"  warn: {CONTROL_SHEET} sheet missing JobNo column")
        return set()
    out: set[int] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        v = row[idx["JobNo"]] if idx["JobNo"] < len(row) else None
        if v is None or str(v).strip() == "":
            continue
        try:
            out.add(int(v))
        except (TypeError, ValueError):
            m = re.search(r"\d+", str(v))
            if m:
                out.add(int(m.group(0)))
            else:
                print(f"  warn: ignoring non-numeric ACTIVE_SHUTDOWNS JobNo {v!r}")
    return out


def _read_trades(wb) -> dict[Any, str]:
    ws = wb[TRADE_SHEET]
    _, idx = _headers(ws)
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not any(row):
            continue
        tid = row[idx["TradeId"]]
        name = _clean_text(row[idx["Trade"]]) or "Unknown"
        out[tid] = ROLE_RENAME.get(name, name)
    return out


def _read_planning(wb, trades: dict[Any, str]) -> dict[int, dict[str, dict[str, int]]]:
    ws = wb[JOB_PLANNING_SHEET]
    _, idx = _headers(ws)
    out: dict[int, dict[str, dict[str, int]]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not any(row):
            continue
        job = row[idx["JobNo"]]
        if job is None:
            continue
        try:
            job_no = int(job)
        except (TypeError, ValueError):
            m = re.search(r"\d+", str(job))
            if not m:
                continue
            job_no = int(m.group(0))
        trade = trades.get(row[idx["CompetencyId"]], "Unknown")
        bucket = out.setdefault(job_no, {})
        cell = bucket.setdefault(trade, {"required": 0, "filled": 0})
        cell["required"] += int(row[idx.get("Required", -1)] or 0) if "Required" in idx else 0
        cell["filled"] += int(row[idx.get("Filled", -1)] or 0) if "Filled" in idx else 0
    return out


def _read_personnel(wb) -> dict[Any, dict[str, str]]:
    ws = wb[PERSONNEL_SHEET]
    _, idx = _headers(ws)
    out: dict[Any, dict[str, str]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not any(row):
            continue
        pid = row[idx["Personnel Id"]]
        if not pid:
            continue
        first = _clean_text(row[idx.get("Given Names", -1)])
        last = _clean_text(row[idx.get("Surname", -1)])
        out[pid] = {
            "name": f"{first} {last}".strip() or "Unknown",
            "role": _clean_text(row[idx.get("Primary Role", -1)]) or "Unknown",
            "mobile": _mobile(row[idx.get("Mobile", -1)] if "Mobile" in idx else None),
            "hire_company": _clean_text(row[idx.get("Hire Company", -1)]),
        }
    return out


def _read_roster(wb, active_jobnos: set[int]) -> dict[int, dict[str, Any]]:
    ws = wb[ROSTER_VIEW_SHEET]
    _, idx = _headers(ws)
    out: dict[int, dict[str, Any]] = {
        j: {"client": None, "site": None, "dates": [], "workers": defaultdict(lambda: {
            "dates": [], "sched_types": Counter(), "is_on_location": False,
        })}
        for j in active_jobnos
    }
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not any(row):
            continue
        job = row[idx["Job No"]]
        if job is None:
            continue
        try:
            job = int(job)
        except (TypeError, ValueError):
            m = re.search(r"\d+", str(job))
            if not m:
                continue
            job = int(m.group(0))
        if job not in out:
            continue
        bucket = out[job]
        bucket["client"] = bucket["client"] or _clean_text(row[idx.get("Client")])
        bucket["site"] = bucket["site"] or _clean_text(row[idx.get("Site")])
        d = _to_date(row[idx["Schedule Date"]]) if "Schedule Date" in idx else None
        if d:
            bucket["dates"].append(d)
        pid = row[idx["Personnel Id"]] if "Personnel Id" in idx else None
        if pid:
            w = bucket["workers"][pid]
            if d:
                w["dates"].append(d)
            sched = row[idx["Schedule Type"]] if "Schedule Type" in idx else None
            if sched:
                w["sched_types"][_clean_text(sched)] += 1
            if "IsOnLocation" in idx and row[idx["IsOnLocation"]]:
                w["is_on_location"] = True
    return out


def _existing_jobnos() -> set[int]:
    out: set[int] = set()
    for name in ("covalent", "tronox", "csbp"):
        path = DATA_DIR / f"{name}.json"
        if not path.exists():
            continue
        data = json.loads(path.read_text())
        for s in data.get("shutdowns", []):
            src = s.get("_source", {}) or {}
            for raw in (
                src.get("rapid_crews_roster_id"),
                src.get("job_no"),
                (src.get("target_source") or {}).get("job_no"),
            ):
                try:
                    out.add(int(raw))
                except (TypeError, ValueError):
                    pass
    return out


def _load_company(company_key: str, company_name: str) -> dict:
    path = DATA_DIR / f"{company_key}.json"
    if path.exists():
        return json.loads(path.read_text())
    return {"company": company_name, "generated_at": None, "shutdowns": []}


def _project_label(base: str, start: dt.date | None, job_no: int) -> str:
    if start:
        suffix = f"{MONTH_NAME[start.month]} {start.year}"
        if base.upper() == "CSBP":
            return f"CSBP Shutdown {suffix}"
        return f"{base} {suffix}"
    return f"{base} Job {job_no}"


def _build_placeholder(job_no: int, planning: dict[str, dict[str, int]], roster_raw: dict[str, Any], personnel: dict[Any, dict[str, str]]) -> tuple[str, str, dict] | None:
    mapped = _map_client_site(roster_raw.get("client"), roster_raw.get("site"), job_no)
    if not mapped:
        return None
    company_key, company_name, dashboard_site, label_base = mapped

    dates = list(roster_raw.get("dates") or [])
    start = min(dates) if dates else dt.date.today()
    end = max(dates) if dates else start

    roster_entries = []
    crew_split: Counter[str] = Counter()
    mobilised_by_role: Counter[str] = Counter()
    labour_hire_split: Counter[str] = Counter()
    scheduled_by_role: Counter[str] = Counter()

    for pid, w in (roster_raw.get("workers") or {}).items():
        person = personnel.get(pid)
        if not person:
            continue
        w_dates = w.get("dates") or dates
        if not w_dates:
            continue
        sched_counter = w.get("sched_types") or Counter()
        dominant_sched = sched_counter.most_common(1)[0][0] if sched_counter else "Scheduled"
        crew = CREW_LABEL.get(dominant_sched, str(dominant_sched))
        role = person.get("role") or "Unknown"
        entry = {
            "name": person.get("name") or "Unknown",
            "role": role,
            "shift": crew,
            "start": min(w_dates).isoformat(),
            "end": max(w_dates).isoformat(),
            "personnel_id": str(pid),
            "tickets": {},
        }
        if person.get("mobile"):
            entry["mobile"] = person["mobile"]
        roster_entries.append(entry)
        crew_split[crew] += 1
        scheduled_by_role[role] += 1
        if w.get("is_on_location"):
            mobilised_by_role[role] += 1
        if person.get("hire_company"):
            labour_hire_split[person["hire_company"]] += 1

    required = {role: int(v.get("required", 0)) for role, v in planning.items() if v.get("required") or v.get("filled")}
    filled = {role: int(v.get("filled", 0)) for role, v in planning.items() if v.get("required") or v.get("filled")}

    for role in scheduled_by_role:
        required.setdefault(role, 0)
        filled.setdefault(role, 0)

    shutdown_id = f"{company_key}-{start.isoformat()[:7]}-{job_no}"
    shutdown = {
        "id": shutdown_id,
        "name": _project_label(label_base, start, job_no),
        "site": dashboard_site,
        "start_date": start.isoformat(),
        "end_date": end.isoformat(),
        "status": _status(start, end),
        "required_by_role": required,
        "filled_by_role": filled,
        "crew_split": dict(crew_split),
        "mobilised_by_role": dict(mobilised_by_role),
        "labour_hire_split": dict(labour_hire_split),
        "roster": roster_entries,
        "_source": {
            "job_no": job_no,
            "source_format": "rapidcrews_macro_safety_pass",
            "required_target_source": "RAPID_CREWS_JOB_PLANNING",
            "target_source": {
                "source": "rapid_crews_job_planning_view",
                "job_no": job_no,
                "total_required": sum(required.values()),
                "total_filled": sum(filled.values()),
            },
        },
    }
    return company_key, company_name, shutdown


def main() -> int:
    if not MACRO_FILE.exists():
        print("ensure_active_shutdowns: no macro workbook found; skipped")
        return 0

    wb = openpyxl.load_workbook(MACRO_FILE, data_only=True, read_only=True)
    try:
        active_jobnos = _read_active_jobnos(wb)
        if not active_jobnos:
            print("ensure_active_shutdowns: no active JobNos found; skipped")
            return 0
        trades = _read_trades(wb)
        planning_all = _read_planning(wb, trades)
        personnel = _read_personnel(wb)
        roster = _read_roster(wb, active_jobnos)
    finally:
        wb.close()

    existing = _existing_jobnos()
    missing = sorted(active_jobnos - existing)
    if not missing:
        print(f"ensure_active_shutdowns: all {len(active_jobnos)} active JobNos already emitted")
        return 0

    now = dt.datetime.now(dt.timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")
    emitted = 0
    for job_no in missing:
        planning = planning_all.get(job_no)
        if not planning:
            print(f"  warn: active JobNo {job_no} has no JobPlanningView rows — cannot emit dashboard card")
            continue
        built = _build_placeholder(job_no, planning, roster.get(job_no, {}), personnel)
        if not built:
            continue
        company_key, company_name, shutdown = built
        payload = _load_company(company_key, company_name)
        payload["generated_at"] = now
        payload.setdefault("shutdowns", []).append(shutdown)
        payload["shutdowns"].sort(key=lambda s: (s.get("start_date") or "", s.get("id") or ""))
        out = DATA_DIR / f"{company_key}.json"
        out.write_text(json.dumps(payload, indent=2))

        HISTORY_DIR.mkdir(parents=True, exist_ok=True)
        hist = {
            "company_key": company_key,
            "client_name": company_name,
            "archived_at": now,
            "shutdown": shutdown,
        }
        (HISTORY_DIR / f"{shutdown['id']}.json").write_text(json.dumps(hist, indent=2))
        emitted += 1
        print(f"  emitted missing active JobNo {job_no}: {shutdown['id']} ({sum(shutdown['required_by_role'].values())} required / {sum(shutdown['filled_by_role'].values())} filled)")

    print(f"ensure_active_shutdowns: emitted {emitted} missing active shutdown(s)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
