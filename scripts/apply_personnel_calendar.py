#!/usr/bin/env python3
"""Export RapidCrews personnel calendar availability for dashboard use.

Reads `xpbi02 PersonnelCalendarView` from Rapidcrews Macro Data.xlsx and emits
`data/personnel_calendar.json`. The front-end uses this to mark worker-matrix
shutdown cells with a red cross when a worker is unavailable because they are
booked elsewhere with SRG or have booked time off.

The sheet has only recently been added, so this parser is deliberately tolerant
of column naming. It searches for common header names and skips rows that do not
have a usable person + date range.
"""
from __future__ import annotations

import datetime as dt
import json
import pathlib
import re
import sys
from typing import Any

import openpyxl

REPO_ROOT = pathlib.Path(__file__).resolve().parent.parent
DATA_DIR = REPO_ROOT / "data"
MACRO_FILE = DATA_DIR / "raw" / "Rapidcrews Macro Data.xlsx"
CALENDAR_SHEET = "xpbi02 PersonnelCalendarView"
OUT_FILE = DATA_DIR / "personnel_calendar.json"


def _clean(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\xa0", " ")).strip()


def _name_key(*parts: Any) -> str:
    text = " ".join(_clean(p) for p in parts if _clean(p))
    return re.sub(r"[^a-z]+", "", text.lower())


def _headers(ws) -> dict[str, int]:
    try:
        header = list(next(ws.iter_rows(max_row=1, values_only=True)))
    except StopIteration:
        return {}
    return {_clean(h): i for i, h in enumerate(header) if _clean(h)}


def _find_idx(headers: dict[str, int], *needles: str) -> int | None:
    lowered = {k.lower(): v for k, v in headers.items()}
    for needle in needles:
        n = needle.lower()
        if n in lowered:
            return lowered[n]
    for key, idx in lowered.items():
        for needle in needles:
            n = needle.lower()
            if n and n in key:
                return idx
    return None


def _get(row: tuple[Any, ...], idx: int | None) -> Any:
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _date(value: Any) -> str:
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    text = _clean(value)
    if not text:
        return ""
    # Already ISO-like.
    m = re.search(r"(\d{4})[-/](\d{1,2})[-/](\d{1,2})", text)
    if m:
        y, mo, d = map(int, m.groups())
        try:
            return dt.date(y, mo, d).isoformat()
        except ValueError:
            return ""
    # Australian date-like.
    m = re.search(r"(\d{1,2})[-/](\d{1,2})[-/](\d{2,4})", text)
    if m:
        d, mo, y = map(int, m.groups())
        if y < 100:
            y += 2000
        try:
            return dt.date(y, mo, d).isoformat()
        except ValueError:
            return ""
    return ""


def _event_type(*parts: Any) -> str:
    text = " | ".join(_clean(p) for p in parts if _clean(p))
    lowered = text.lower()
    if any(term in lowered for term in ("leave", "rdo", "annual", "personal", "sick", "time off", "unavailable", "booked off", "off")):
        return "time_off"
    if any(term in lowered for term in ("job", "roster", "shutdown", "project", "work", "booked", "assignment", "site")):
        return "srg_work"
    return "calendar_booking" if text else "calendar_booking"


def _read_calendar() -> list[dict[str, Any]]:
    if not MACRO_FILE.exists():
        print("apply_personnel_calendar: macro workbook not found; skipped")
        return []

    try:
        wb = openpyxl.load_workbook(MACRO_FILE, data_only=True, read_only=True)
    except Exception as exc:
        print(f"::warning::apply_personnel_calendar could not open workbook: {exc}", file=sys.stderr)
        return []

    events: list[dict[str, Any]] = []
    try:
        if CALENDAR_SHEET not in wb.sheetnames:
            print(f"apply_personnel_calendar: sheet not found: {CALENDAR_SHEET}; skipped")
            return []

        ws = wb[CALENDAR_SHEET]
        headers = _headers(ws)
        if not headers:
            print("apply_personnel_calendar: no headers found; skipped")
            return []

        pid_i = _find_idx(headers, "Personnel Id", "PersonnelId", "Personnel ID", "Employee Id", "EmployeeID")
        first_i = _find_idx(headers, "Given Names", "Given Name", "First Name", "Forename")
        last_i = _find_idx(headers, "Surname", "Last Name", "Family Name")
        full_i = _find_idx(headers, "Personnel", "Employee", "Employee Name", "Name", "Full Name", "Resource")
        role_i = _find_idx(headers, "Role", "Trade", "Position", "Discipline")
        start_i = _find_idx(headers, "Start Date", "Start", "From Date", "From", "Date From", "Booked From")
        end_i = _find_idx(headers, "End Date", "End", "To Date", "To", "Date To", "Booked To", "Finish")
        desc_i = _find_idx(headers, "Description", "Calendar", "Activity", "Event", "Reason", "Leave Type", "Booking Type", "Status")
        company_i = _find_idx(headers, "Company", "Client", "Customer", "Site", "Project", "Job", "Job Description")
        job_i = _find_idx(headers, "Job No", "JobNo", "Job", "Roster Id", "RosterId")

        for row_no, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                if not row or not any(row):
                    continue

                first = _clean(_get(row, first_i))
                last = _clean(_get(row, last_i))
                full = _clean(_get(row, full_i))
                name = _clean(" ".join(p for p in (first, last) if p)) or full
                key = _name_key(name)
                if not key:
                    continue

                start = _date(_get(row, start_i))
                end = _date(_get(row, end_i)) or start
                if not start:
                    continue

                desc = _clean(_get(row, desc_i))
                company = _clean(_get(row, company_i))
                job = _clean(_get(row, job_i))
                role = _clean(_get(row, role_i))
                event_type = _event_type(desc, company, job)

                events.append({
                    "personnel_id": _clean(_get(row, pid_i)),
                    "name": name,
                    "name_key": key,
                    "role": role,
                    "start": start,
                    "end": end,
                    "type": event_type,
                    "description": desc,
                    "company_or_job": company,
                    "job_no": job,
                    "source_row": row_no,
                })
            except Exception as exc:
                print(f"::warning::apply_personnel_calendar skipped row {row_no}: {exc}", file=sys.stderr)
    finally:
        wb.close()

    return events


def main() -> int:
    events = _read_calendar()
    payload = {
        "generated_at": dt.datetime.utcnow().replace(microsecond=0).isoformat() + "Z",
        "source": str(MACRO_FILE.relative_to(REPO_ROOT)),
        "sheet": CALENDAR_SHEET,
        "events": events,
    }
    OUT_FILE.write_text(json.dumps(payload, indent=2))
    print(f"apply_personnel_calendar: wrote {len(events)} events to {OUT_FILE.relative_to(REPO_ROOT)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
